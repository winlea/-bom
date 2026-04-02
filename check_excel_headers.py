from openpyxl import load_workbook


def check_excel_headers():
    """使用openpyxl检查Excel文件的实际表头内容"""
    print("开始检查Excel文件的实际表头内容...")
    print("=" * 60)

    # Excel文件路径
    excel_file = "E:\\NEW\\-bom\\WH-DFYF-H47-26-00 H47产品履历模版2026-2-7.xlsx"

    print(f"检查Excel文件: {excel_file}")
    print()

    try:
        # 加载工作簿
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        print(f"工作表名称: {ws.title}")
        print()

        # 读取前10行，寻找表头
        print("前10行的内容:")
        for row_idx in range(1, 11):
            row = ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
            row_values = list(row)[0]
            # 过滤掉空值
            non_empty_values = [v for v in row_values if v is not None]
            if non_empty_values:
                print(f"行 {row_idx}: {non_empty_values}")

        print()
        print("=" * 60)
        print("检查完成!")

    except Exception as e:
        print(f"错误: {e}")


if __name__ == "__main__":
    check_excel_headers()
