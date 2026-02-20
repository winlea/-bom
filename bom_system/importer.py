import base64
import csv
from io import BytesIO, TextIOWrapper
from typing import Any, Dict, List

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import BomTable, db
from .services import _validate_image_bytes, save_base64_image, save_url_image

# 网络请求默认超时（秒）
HTTP_FETCH_TIMEOUT = 8

REQUIRED_COLS = {"part_number"}
OPTIONAL_COLS = {
    "part_name",
    "image_url",
    "image_base64",
    "assembly_level",
    "sequence",
    "drawing_2d",
    "drawing_3d",
    "original_material",
    "final_material_cn",
    "part_category",
    "net_weight_kg",
}

# Chinese header synonyms mapping to internal keys
HEADER_SYNONYMS = {
    "产品编号": "part_number",
    "物料编号": "part_number",
    "图号": "part_number",
    "零件名称": "part_name",
    "产品名称": "part_name",
    "名称": "part_name",
    "装配等级": "assembly_level",
    "层级": "assembly_level",
    "序号": "sequence",
    "2D图号": "drawing_2d",
    "3D图号": "drawing_3d",
    "零件简图": "image_embed",
    "图片": "image_embed",
    "原图材料": "original_material",
    "终审拟代材料（中国标准）": "final_material_cn",
    "终审拟代材料(中国标准)": "final_material_cn",
    "零件分类": "part_category",
    "产品净重KG/PCS": "net_weight_kg",
    "净重": "net_weight_kg",
}


def detect_level_and_part_from_b_to_h(row_values) -> tuple[int | None, str | None]:
    """Fallback: detect assembly level and part number by scanning columns B..H (2..8).
    Returns (level, part_number). Level is 1..7 mapping to B..H if a non-empty string found.
    """
    if not row_values:
        return None, None
    for idx in range(1, 8):  # 0-based: 1..7 -> Excel B..H
        val = row_values[idx] if idx < len(row_values) else None
        if isinstance(val, str):
            s = val.strip()
            if s:
                return idx, s  # level = idx (B=1)
        elif val is not None:
            # numeric-like values also acceptable
            return idx, str(val)
    return None, None


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def import_csv(file_bytes: bytes, project_id: int | None = None) -> Dict[str, Any]:
    """
    兼容性更强的 CSV 导入：
    - 尝试解析 image_base64 / image_url，但图片抓取失败不阻止记录创建（将错误记录在 errors 中）
    - 直接使用数据库模型创建 BomTable，避免依赖外部 save_url_image 在网络异常时抛出并中断整个导入
    """
    f = TextIOWrapper(BytesIO(file_bytes), encoding="utf-8")
    reader = csv.DictReader(f)
    
    # 构建表头映射，支持中文表头
    header_map: Dict[str, str] = {}
    for h in reader.fieldnames or []:
        # 优先使用同义词映射
        mapped_key = HEADER_SYNONYMS.get(h)
        if mapped_key:
            header_map[mapped_key] = h
        else:
            # 否则使用标准化表头
            normalized = _normalize_header(h)
            header_map[normalized] = h
    
    if "part_number" not in header_map:
        raise ValueError("未找到 '产品编号/part_number' 列，请检查表头")
    
    # 为了保持兼容性，仍然构建原始的 headers 字典
    headers = {k: header_map[k] for k in header_map}

    created = 0
    errors: List[str] = []
    for i, row in enumerate(reader, start=2):  # header at 1
        pn = (row.get(headers.get("part_number", "")) or "").strip()
        part_name = (row.get(headers.get("part_name", ""), "") or "").strip() or None
        image_url = (row.get(headers.get("image_url", ""), "") or "").strip()
        image_base64 = (row.get(headers.get("image_base64", ""), "") or "").strip()

        if not pn:
            # skip rows without part number
            continue

        img_bytes = None
        img_url_to_save = None

        # 优先处理 base64 内嵌图
        try:
            if image_base64:
                enc = (
                    image_base64.split(",", 1)[1]
                    if image_base64.lower().startswith("data:") and "," in image_base64
                    else image_base64
                )
                try:
                    img_try = base64.b64decode(enc, validate=True)
                    ok, _ = _validate_image_bytes(img_try)
                    if ok:
                        img_bytes = img_try
                except Exception as e:
                    # base64 解码失败，记录错误但不阻止创建
                    errors.append(f"row {i}: invalid base64 image ({e})")
                    img_bytes = None
            elif image_url:
                try:
                    r = requests.get(image_url, timeout=HTTP_FETCH_TIMEOUT)
                    if r.status_code == 200 and r.content:
                        ok, _ = _validate_image_bytes(r.content)
                        if ok:
                            img_bytes = r.content
                            img_url_to_save = image_url
                        else:
                            # 非法图片字节，仍保存 URL 以便后续人工查看
                            img_url_to_save = image_url
                            errors.append(f"row {i}: fetched content not a valid image")
                    else:
                        # 无法成功下载，记录为 warning，但继续创建记录并保留 image_url
                        img_url_to_save = image_url
                        errors.append(
                            f"row {i}: failed to fetch image (status {r.status_code})"
                        )
                except Exception as e:
                    # 网络/SSL 等异常不阻止创建，记录错误并保留 image_url
                    img_url_to_save = image_url
                    errors.append(f"row {i}: Failed to fetch URL: {e}")
        except Exception as e:
            # 防御性捕获，记录后继续
            errors.append(f"row {i}: image handling unexpected error: {e}")

        # 创建 DB 记录（即使图片处理有问题也创建）
        rec = BomTable(
            project_id=project_id,
            part_number=pn,
            part_name=part_name,
            sequence=None,
            assembly_level=None,
            bom_sort=0,
            drawing_2d=None,
            drawing_3d=None,
            original_material=None,
            final_material_cn=None,
            part_category=None,
            net_weight_kg=None,
            image_data=img_bytes,
            image_url=img_url_to_save,
        )

        db.session.add(rec)
        try:
            db.session.commit()
            created += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"row {i}: db commit failed: {e}")

    return {"created": created, "errors": errors}


def _find_header_row(ws: Worksheet, max_scan: int = 50) -> int:
    """Scan first max_scan rows to find a header row that contains synonyms for part_number or part_name."""
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
    ):
        texts = [str(v or "").strip() for v in row]
        norm = [_normalize_header(t) for t in texts]
        if any(t in HEADER_SYNONYMS for t in texts) or any(
            n in ("part_number", "part_name", "image_url", "image_base64") for n in norm
        ):
            return r_idx
    return 1


def import_xlsx(file_bytes: bytes, project_id: int | None = None) -> Dict[str, Any]:
    print(f"DEBUG: import_xlsx called with project_id: {project_id}")
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    headers = [
        str(v or "").strip()
        for v in next(
            ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
        )
    ]

    # Build header_map using synonyms
    header_map: Dict[str, int] = {}
    for idx, label in enumerate(headers):
        key = HEADER_SYNONYMS.get(label)
        if key:
            header_map[key] = idx
        else:
            n = _normalize_header(label)
            if n in (REQUIRED_COLS | OPTIONAL_COLS):
                header_map[n] = idx

    if "part_number" not in header_map:
        raise ValueError("未找到 '产品编号/part_number' 列，请检查表头")

    # Build map of embedded images by row index (1-based row number)
    img_by_row: Dict[int, bytes] = {}
    imgs = getattr(ws, "_images", [])
    for img in imgs:
        try:
            anc = getattr(img, "anchor", None)
            row = None
            if hasattr(anc, "row"):
                row = anc.row + 1
            elif hasattr(anc, "_from") and hasattr(anc._from, "row"):
                row = anc._from.row + 1
            if row:
                data = img._data()
                ok, _ = _validate_image_bytes(data)
                if ok:
                    img_by_row[row] = data
        except Exception:
            continue

    created = 0
    errors: List[str] = []

    bom_sort_counter = 0
    last_seq_by_level: Dict[int, str] = {}
    used_sequences: set[str] = set()
    top_level_seen_parts: set[str] = set()

    def ensure_unique(seq: str) -> str:
        if seq not in used_sequences:
            return seq
        parts = seq.split(".")
        try:
            last = int(parts[-1])
        except Exception:
            last = 0
        while True:
            last += 1
            cand = ".".join(parts[:-1] + [str(last)])
            if cand not in used_sequences:
                return cand

    for i, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        row_values = list(row) if row is not None else []

        def get_col(name: str):
            idx = header_map.get(name)
            return (row[idx] if idx is not None else None) if row is not None else None

        pn = str(get_col("part_number") or "").strip()
        part_name = str(get_col("part_name") or "").strip() or None

        # extra columns
        assembly_level_val = get_col("assembly_level")
        try:
            assembly_level = (
                int(assembly_level_val)
                if assembly_level_val not in (None, "")
                else None
            )
        except Exception:
            assembly_level = None
        sequence_val = str(get_col("sequence") or "").strip()
        drawing_2d = str(get_col("drawing_2d") or "").strip() or None
        drawing_3d = str(get_col("drawing_3d") or "").strip() or None
        original_material = str(get_col("original_material") or "").strip() or None
        final_material_cn = str(get_col("final_material_cn") or "").strip() or None
        part_category = str(get_col("part_category") or "").strip() or None
        net_weight_raw = get_col("net_weight_kg")
        try:
            net_weight_kg = (
                float(net_weight_raw) if net_weight_raw not in (None, "") else None
            )
        except Exception:
            net_weight_kg = None

        # Fallback: detect level and part number from B..H columns if missing
        if (not assembly_level) or (not pn):
            lvl_from_bh, pn_from_bh = detect_level_and_part_from_b_to_h(row_values)
            if not assembly_level and lvl_from_bh:
                assembly_level = int(lvl_from_bh)
            if not pn and pn_from_bh:
                pn = pn_from_bh.strip()

        # Skip if still no part number
        if not pn:
            continue

        # Top-level duplicate control
        if (assembly_level or 1) == 1:
            if pn in top_level_seen_parts:
                # skip duplicate top-level
                continue
            top_level_seen_parts.add(pn)

        # fill sequence according to assembly_level hierarchy
        if sequence_val:
            sequence = sequence_val
            last_seq_by_level[assembly_level or 1] = sequence
            used_sequences.add(sequence)
        else:
            lvl = assembly_level or 1
            if lvl == 1:
                prev = last_seq_by_level.get(1)
                if prev:
                    try:
                        num = int(prev.split(".")[0]) + 1
                    except Exception:
                        num = 1
                    sequence = str(num)
                else:
                    sequence = "1"
                sequence = ensure_unique(sequence)
                last_seq_by_level[1] = sequence
                used_sequences.add(sequence)
            else:
                parent_lvl = lvl - 1
                parent_seq = last_seq_by_level.get(parent_lvl)
                if not parent_seq:
                    # synthesize missing parent chain up to lvl-1
                    # create a new top-level if absent
                    top_prev = last_seq_by_level.get(1)
                    base_top = str(int(top_prev.split(".")[0]) + 1) if top_prev else "1"
                    base_top = ensure_unique(base_top)
                    last_seq_by_level[1] = base_top
                    used_sequences.add(base_top)
                    # descend to parent level
                    for x in range(2, lvl):
                        new_child = last_seq_by_level.get(x)
                        if not new_child:
                            parent = last_seq_by_level[x - 1]
                            new_child = parent + ".1"
                            new_child = ensure_unique(new_child)
                            last_seq_by_level[x] = new_child
                            used_sequences.add(new_child)
                    parent_seq = last_seq_by_level.get(parent_lvl)
                sibling_prev = last_seq_by_level.get(lvl)
                if sibling_prev and sibling_prev.startswith(parent_seq + "."):
                    try:
                        last_idx = int(sibling_prev.split(".")[-1])
                    except Exception:
                        last_idx = 0
                    sequence = parent_seq + "." + str(last_idx + 1)
                else:
                    sequence = parent_seq + ".1"
                sequence = ensure_unique(sequence)
                last_seq_by_level[lvl] = sequence
                used_sequences.add(sequence)

        # images
        image_url = str(get_col("image_url") or "").strip()
        image_base64 = str(get_col("image_base64") or "").strip()

        bom_sort_counter += 1

        # Prepare image bytes for this record (attach to same row)
        img_bytes = None
        img_url_to_save = None
        try:
            if image_base64:
                # decode base64 data URI
                enc = (
                    image_base64.split(",", 1)[1]
                    if image_base64.lower().startswith("data:") and "," in image_base64
                    else image_base64
                )
                img_bytes_try = base64.b64decode(enc, validate=True)
                ok, _ = _validate_image_bytes(img_bytes_try)
                if ok:
                    img_bytes = img_bytes_try
            elif image_url:
                try:
                    r = requests.get(image_url, timeout=HTTP_FETCH_TIMEOUT)
                    if r.status_code == 200 and r.content:
                        ok, _ = _validate_image_bytes(r.content)
                        if ok:
                            img_bytes = r.content
                            img_url_to_save = image_url
                        else:
                            img_url_to_save = image_url
                    else:
                        img_url_to_save = image_url
                except Exception as e:
                    # 非致命：记录警告并保留 image_url，继续创建记录
                    img_url_to_save = image_url
                    errors.append(f"row {i}: Failed to fetch URL: {e}")
            elif i in img_by_row:
                img_bytes_try = img_by_row[i]
                ok, _ = _validate_image_bytes(img_bytes_try)
                if ok:
                    img_bytes = img_bytes_try
        except Exception as e:
            # 防御性地记录并继续
            errors.append(f"row {i}: image error: {e}")

        # Create record (single row per BOM line)
        rec = BomTable(
            project_id=project_id,
            part_number=pn,
            part_name=part_name,
            sequence=sequence,
            assembly_level=assembly_level,
            bom_sort=bom_sort_counter,
            drawing_2d=drawing_2d,
            drawing_3d=drawing_3d,
            original_material=original_material,
            final_material_cn=final_material_cn,
            part_category=part_category,
            net_weight_kg=net_weight_kg,
            image_data=img_bytes,
            image_url=img_url_to_save,
        )

        db.session.add(rec)
        try:
            db.session.commit()
            created += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f"row {i}: {e}")

    return {"created": created, "errors": errors}
