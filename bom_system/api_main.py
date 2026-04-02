import os
from io import BytesIO
from datetime import datetime

from flask import Blueprint, request, send_file
from openpyxl import load_workbook

from .importer import import_csv, import_xlsx
from .models import BomTable, ImportLog, Project, DrawingChange, db
from .api.response import APIResponse
from .services import (
    get_effective_image_bytes,
    get_image_bytes,
    save_base64_image,
    save_url_image,
)
from .templates.dynamic_ods_generator import (
    DynamicODSGenerator,
    create_simple_ods_from_data,
)

api_bp = Blueprint("api", __name__)


@api_bp.route("/upload/base64", methods=["POST"])
def upload_base64():
    payload = request.get_json(silent=True) or {}
    part_number = payload.get("part_number")
    image_data = payload.get("image_data")
    part_name = payload.get("part_name")
    if not part_number or not image_data:
        return APIResponse.bad_request(
            message="part_number and image_data are required"
        )
    try:
        rec_id = save_base64_image(part_number, image_data, part_name)
    except ValueError as e:
        return APIResponse.bad_request(
            message=str(e)
        )
    return APIResponse.success(
        data={"status": "success", "id": rec_id}
    )


@api_bp.route("/upload/url", methods=["POST"])
def upload_url():
    payload = request.get_json(silent=True) or {}
    part_number = payload.get("part_number")
    url = payload.get("url")
    part_name = payload.get("part_name")
    if not part_number or not url:
        return APIResponse.bad_request(
            message="part_number and url are required"
        )
    try:
        rec_id = save_url_image(part_number, url, part_name)
    except ValueError as e:
        return APIResponse.bad_request(
            message=str(e)
        )
    return APIResponse.success(
        data={"status": "success", "id": rec_id}
    )


@api_bp.route("/download/<int:record_id>", methods=["GET"])
def download_image(record_id: int):
    # attempt exact image; if not present, try ancestor fallback
    try:
        data, mime = get_image_bytes(record_id)
    except LookupError:
        try:
            data, mime = get_effective_image_bytes(record_id)
        except LookupError:
            return APIResponse.not_found(
                message="Image not found"
            )
    return send_file(
        BytesIO(data),
        mimetype=mime,
        as_attachment=False,
        download_name=f"image_{record_id}",
    )


@api_bp.route("/import/bom", methods=["POST"])
def import_bom():
    """Import BOM into a project. multipart/form-data with fields: file, project_id."""
    try:
        if "file" not in request.files:
            return APIResponse.bad_request(
                message="请上传CSV或Excel文件",
                errors={"error": "file is required"}
            )

        try:
            project_id = (
                int(request.form.get("project_id"))
                if request.form.get("project_id")
                else None
            )
        except ValueError:
            return APIResponse.bad_request(
                message="项目ID必须是数字",
                errors={"error": "invalid project_id"}
            )

        f = request.files["file"]
        data = f.read()
        if not data:
            return APIResponse.bad_request(
                message="上传的文件为空",
                errors={"error": "empty file"}
            )

        filename = (f.filename or "").lower()
        try:
            if filename.endswith(".csv"):
                result = import_csv(data, project_id=project_id)
            elif filename.endswith(".xlsx") or filename.endswith(".xlsm"):
                result = import_xlsx(data, project_id=project_id)
            else:
                return APIResponse.bad_request(
                    message="仅支持CSV、XLSX和XLSM文件格式",
                    errors={"error": "Unsupported file type"}
                )
        except ValueError as e:
            return APIResponse.bad_request(
                message="文件格式错误或缺少必要字段",
                errors={"error": str(e)}
            )
        except Exception as e:
            return APIResponse.internal_server_error(
                message=f"导入过程中发生错误: {str(e)}",
                errors={"error": "import failed"}
            )

        # log import result and update project status
        try:
            log = ImportLog(
                project_id=project_id,
                filename=(f.filename or ""),
                created_count=result.get("created", 0),
                errors_count=len(result.get("errors", []) or []),
                message="; ".join(result.get("errors", []) or []),
            )
            db.add(log)
            if project_id:
                proj = Project.query.get(project_id)
                if proj and proj.status == "created" and result.get("created", 0) > 0:
                    proj.status = "imported"
            db.commit()
        except Exception as e:
            # 记录日志失败不影响导入结果
            print(f"记录导入日志失败: {str(e)}")

        return APIResponse.success(
            data={"status": "success", **result}
        )
    except Exception as e:
        return APIResponse.internal_server_error(
            message=f"服务器内部错误: {str(e)}",
            errors={"error": "internal server error"}
        )


@api_bp.route("/items", methods=["GET"])
def list_items():
    """List recent items for UI: supports ?q=part&limit=20"""
    q = (request.args.get("q") or "").strip()
    try:
        limit = int(request.args.get("limit", "20"))
    except ValueError:
        limit = 20
    limit = max(1, min(100, limit))

    query = BomTable.query
    if q:
        like = f"%{q}%"
        query = query.filter(BomTable.part_number.like(like))
    # List all items (including those without images)
    rows = query.order_by(BomTable.created_at.desc()).limit(limit).all()

    def ser(r: BomTable):
        return {
            "id": r.id,
            "part_number": r.part_number,
            "part_name": r.part_name,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }

    return APIResponse.success(
        data={"items": [ser(r) for r in rows]}
    )


@api_bp.route("/parts", methods=["GET"])
def list_parts():
    """List parts with project filter: supports ?project_id=1&q=part"""
    try:
        project_id = (
            int(request.args.get("project_id"))
            if request.args.get("project_id")
            else None
        )
    except ValueError:
        project_id = None

    q = (request.args.get("q") or "").strip()

    # 打印查询参数
    print(f"查询参数: project_id={project_id}, q={q}")

    # 打印BomTable的表名
    print(f"BomTable表名: {BomTable.__tablename__}")

    # 直接执行SQL查询
    from sqlalchemy import text
    with db.session.begin():
        result = db.session.execute(text("SELECT * FROM bom_table WHERE project_id = :project_id LIMIT 5"), {"project_id": project_id})
        rows = result.fetchall()
        print(f"直接SQL查询结果: {len(rows)} 条记录")
        for row in rows:
            print(f"记录: id={row.id}, part_number={row.part_number}, project_id={row.project_id}")

    query = BomTable.query

    # Filter by project_id if provided
    if project_id:
        print(f"按project_id={project_id}过滤")
        query = query.filter(BomTable.project_id == project_id)

    # Filter by search term if provided
    if q:
        like = f"%{q}%"
        query = query.filter(BomTable.part_number.like(like))

    # Get all matching parts
    # 先不排序，直接查询
    rows = query.all()
    print(f"查询到 {len(rows)} 条记录（未排序）")
    
    # 尝试排序
    try:
        rows_sorted = query.order_by(BomTable.bom_sort.asc(), BomTable.created_at.desc()).all()
        print(f"排序后查询到 {len(rows_sorted)} 条记录")
        rows = rows_sorted
    except Exception as e:
        print(f"排序失败: {str(e)}")
        # 排序失败时使用未排序的结果
        pass

    # 打印原始查询结果
    print(f"原始查询结果: {len(rows)} 条记录")
    for i, part in enumerate(rows[:5]):
        print(f"  记录{i+1}: id={part.id}, part_number={part.part_number}, project_id={part.project_id}, sequence={part.sequence}, created_at={part.created_at}")

    # 去重：按project_id、part_number和sequence组合分组，只保留最新的记录
    # 序号不同的零件视为不同的零件
    unique_parts = {}
    for part in rows:
        # 使用(project_id, part_number, sequence)作为键，确保不同序号的零件不被去重
        # 处理sequence为None的情况
        sequence = part.sequence or ""
        key = (part.project_id, part.part_number, sequence)
        if key not in unique_parts:
            unique_parts[key] = part
        elif part.created_at and unique_parts[key].created_at and part.created_at > unique_parts[key].created_at:
            unique_parts[key] = part

    # 转换为列表
    unique_rows = list(unique_parts.values())
    print(f"去重后结果: {len(unique_rows)} 条记录")
    for i, part in enumerate(unique_rows[:5]):
        print(f"  去重后记录{i+1}: id={part.id}, part_number={part.part_number}, project_id={part.project_id}, sequence={part.sequence}")

    def ser_part(r: BomTable):
        # 构建图片URL
        image_url = r.image_url
        if not image_url and r.image_data:
            # 如果有图片数据但没有URL，使用图片下载端点
            image_url = f"/api/parts/{r.id}/image"

        return {
            "id": r.id,
            "project_id": r.project_id,
            "part_number": r.part_number,
            "part_name": r.part_name,
            "sequence": r.sequence,
            "assembly_level": r.assembly_level,
            "original_material": r.original_material,
            "final_material_cn": r.final_material_cn,
            "net_weight_kg": r.net_weight_kg,
            "drawing_2d": r.drawing_2d,
            "drawing_3d": r.drawing_3d,
            "image_url": image_url,
            "part_category": r.part_category,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "has_image": bool(r.image_data)
        }

    return APIResponse.success(
        data={"items": [ser_part(r) for r in unique_rows]}
    )


@api_bp.route("/parts/<int:part_id>/image", methods=["GET"])
def get_part_image(part_id):
    """获取零件图片"""
    try:
        part = BomTable.query.get(part_id)
        if not part:
            return APIResponse.not_found(
                message="Part not found"
            )

        if not part.image_data:
            return APIResponse.not_found(
                message="No image data"
            )

        # 尝试检测图片类型
        try:
            import imghdr
            image_type = imghdr.what(None, part.image_data)
            if not image_type:
                image_type = "png"  # 默认类型
        except Exception as img_err:
            print(f"图片类型检测错误: {img_err}")
            image_type = "png"  # 出错时使用默认类型

        # 尝试创建BytesIO对象
        try:
            from io import BytesIO
            img_data = BytesIO(part.image_data)
        except Exception as io_err:
            print(f"BytesIO创建错误: {io_err}")
            return APIResponse.internal_server_error(
                message=f"Image data error: {str(io_err)}",
                errors={"error": "image data error"}
            )

        return send_file(
            img_data,
            mimetype=f"image/{image_type}",
            as_attachment=False,
            download_name=f"part_{part_id}.{image_type}",
        )
    except Exception as e:
        print(f"获取图片失败: {e}")
        return APIResponse.internal_server_error(
            message=f"Get part image failed: {str(e)}",
            errors={"error": "get part image failed"}
        )


@api_bp.route("/parts", methods=["POST"])
def create_part():
    """Create a new part"""
    try:
        data = request.get_json()
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        # Required fields
        part_number = data.get("part_number")
        part_name = data.get("part_name")
        project_id = data.get("project_id")

        if not part_number or not part_name or not project_id:
            return APIResponse.bad_request(
                message="part_number, part_name, and project_id are required"
            )

        # Create new part
        part = BomTable(
            part_number=part_number,
            part_name=part_name,
            project_id=project_id,
            sequence=data.get("sequence"),
            assembly_level=data.get("assembly_level"),
            original_material=data.get("original_material"),
            final_material_cn=data.get("final_material_cn"),
            net_weight_kg=data.get("net_weight_kg"),
            drawing_2d=data.get("drawing_2d"),
            drawing_3d=data.get("drawing_3d"),
            image_url=data.get("image_url"),
            part_category=data.get("part_category"),
        )

        db.add(part)
        db.commit()

        return APIResponse.created(
            data={
                "id": part.id,
                "part_number": part.part_number,
                "part_name": part.part_name,
                "project_id": part.project_id,
                "created_at": (
                    part.created_at.isoformat() if part.created_at else None
                ),
            }
        )
    except Exception as e:
        db.rollback()
        return APIResponse.internal_server_error(
            message=f"Create part failed: {str(e)}",
            errors={"error": "create part failed"}
        )


@api_bp.route("/parts/<int:part_id>", methods=["PUT"])
def update_part(part_id):
    """Update an existing part"""
    try:
        data = request.get_json()
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        part = BomTable.query.get(part_id)
        if not part:
            return APIResponse.not_found(
                message="part not found"
            )

        # Update fields
        if "part_name" in data:
            part.part_name = data["part_name"]
        if "sequence" in data:
            part.sequence = data["sequence"]
        if "assembly_level" in data:
            part.assembly_level = data["assembly_level"]
        if "original_material" in data:
            part.original_material = data["original_material"]
        if "final_material_cn" in data:
            part.final_material_cn = data["final_material_cn"]
        if "net_weight_kg" in data:
            part.net_weight_kg = data["net_weight_kg"]
        if "drawing_2d" in data:
            part.drawing_2d = data["drawing_2d"]
        if "drawing_3d" in data:
            part.drawing_3d = data["drawing_3d"]
        if "image_url" in data:
            part.image_url = data["image_url"]
        if "part_category" in data:
            part.part_category = data["part_category"]

        db.commit()

        return APIResponse.success(
            data={
                "id": part.id,
                "part_number": part.part_number,
                "part_name": part.part_name,
                "project_id": part.project_id,
                "created_at": part.created_at.isoformat() if part.created_at else None,
            }
        )
    except Exception as e:
        db.rollback()
        return APIResponse.internal_server_error(
            message=f"Update part failed: {str(e)}",
            errors={"error": "update part failed"}
        )


@api_bp.route("/parts/<int:part_id>", methods=["DELETE"])
def delete_part(part_id):
    """Delete a part"""
    try:
        part = BomTable.query.get(part_id)
        if not part:
            return APIResponse.not_found(
                message="part not found"
            )

        db.delete(part)
        db.commit()

        return APIResponse.success(
            message="part deleted",
            data={"status": "success"}
        )
    except Exception as e:
        db.rollback()
        return APIResponse.internal_server_error(
            message=f"Delete part failed: {str(e)}",
            errors={"error": "delete part failed"}
        )


@api_bp.route("/parts/<int:part_id>", methods=["GET"])
def get_part(part_id):
    """Get a single part"""
    try:
        part = BomTable.query.get(part_id)
        if not part:
            return APIResponse.not_found(
                message="part not found"
            )

        return APIResponse.success(
            data={
                "id": part.id,
                "project_id": part.project_id,
                "part_number": part.part_number,
                "part_name": part.part_name,
                "sequence": part.sequence,
                "assembly_level": part.assembly_level,
                "original_material": part.original_material,
                "final_material_cn": part.final_material_cn,
                "net_weight_kg": part.net_weight_kg,
                "drawing_2d": part.drawing_2d,
                "drawing_3d": part.drawing_3d,
                "image_url": part.image_url,
                "part_category": part.part_category,
                "created_at": part.created_at.isoformat() if part.created_at else None,
            }
        )
    except Exception as e:
        return APIResponse.internal_server_error(
            message=f"Get part failed: {str(e)}",
            errors={"error": "get part failed"}
        )


@api_bp.route("/reports/generate", methods=["POST"])
def generate_report():
    """Generate ODS report for a project"""
    try:
        # 确保请求方法是POST
        if request.method != 'POST':
            return APIResponse.error(
                message="Method not allowed",
                status_code=405
            )

        # 获取请求数据
        data = request.get_json()
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        project_id = data.get("project_id")
        if not project_id:
            return APIResponse.bad_request(
                message="project_id is required"
            )

        # Get project information
        project = Project.query.get(project_id)
        if not project:
            return APIResponse.not_found(
                message="project not found"
            )

        # Get project parts
        parts = (
            BomTable.query.filter_by(project_id=project_id)
            .order_by(BomTable.bom_sort.asc(), BomTable.created_at.desc())
            .all()
        )

        if not parts:
            return APIResponse.not_found(
                message="no parts found for this project"
            )

        # Prepare data for report
        report_data = []
        for part in parts:
            report_data.append(
                {
                    "sequence": part.sequence,
                    "part_number": part.part_number,
                    "part_name": part.part_name,
                    "drawing_2d": part.drawing_2d,
                    "drawing_3d": part.drawing_3d,
                    "original_material": part.original_material,
                    "final_material_cn": part.final_material_cn,
                    "part_category": part.part_category,
                    "net_weight_kg": part.net_weight_kg,
                }
            )

        project_info = {
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "status": project.status,
            "generated_time": (
                project.created_at.isoformat() if project.created_at else None
            ),
        }

        # 直接使用简单ODS创建
        output_file = create_simple_ods_from_data(
            data=report_data, project_info=project_info
        )

        # Send the file
        return send_file(
            output_file,
            as_attachment=True,
            download_name=f"{project.name}_ODS_Report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        print(f"Report generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        return APIResponse.internal_server_error(
            message=f"Generate report failed: {str(e)}",
            errors={"error": "generate report failed"}
        )


@api_bp.route("/psw/generate", methods=["POST"])
def generate_psw():
    """Generate PSW document for a part"""
    try:
        # 确保请求方法是POST
        if request.method != 'POST':
            return APIResponse.error(
                message="Method not allowed",
                status_code=405
            )

        # 获取请求数据
        data = request.get_json()
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        part_id = data.get("part_id")
        if not part_id:
            return APIResponse.bad_request(
                message="part_id is required"
            )

        # Get part information
        part = BomTable.query.get(part_id)
        if not part:
            return APIResponse.not_found(
                message="part not found"
            )

        # Get project information
        project = None
        if part.project_id:
            project = Project.query.get(part.project_id)
            if not project:
                print(f"Project not found for part {part_id}")

        # 模板文件路径 - 直接使用新模板文件
        template_file = r"C:\Users\Administrator\univer\PSW_Template.xlsx"
        if not os.path.exists(template_file):
            return APIResponse.internal_server_error(
                message="PSW template file not found"
            )

        # 当前日期
        current_date = datetime.now().strftime("%m-%d-%y")

        # 使用pywin32库来操作Excel文件，确保格式完全一致
        import win32com.client
        import pythoncom
        import tempfile
        import shutil

        # 初始化COM环境
        pythoncom.CoInitialize()

        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file_path = temp_file.name

        # 复制模板文件到临时文件
        shutil.copyfile(template_file, temp_file_path)

        # 启动Excel应用程序
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            # 打开临时文件
            wb = excel.Workbooks.Open(temp_file_path)
            # 选择PSW工作表
            ws = wb.Worksheets('PSW')

            # 查询最新的图纸变更信息
            drawing_change = DrawingChange.query.filter_by(part_id=part_id).order_by(DrawingChange.created_at.desc()).first()
            
            # 填充核心字段
            fields_to_fill = {
                (5, 3): part.part_name if part.part_name else "N/A",  # 零件名称
                (5, 8): part.part_number if part.part_number else "N/A",  # 零件号
                (5, 13): part.drawing_2d if part.drawing_2d else "N/A",  # 图纸编号
                (7, 8): current_date,  # 日期1
                (7, 13): current_date,  # 日期2
                (9, 13): part.net_weight_kg if part.net_weight_kg is not None else 0,  # 重量，当为None时使用0
                (46, 4): "冲压/弧焊/电泳",  # 生产工艺
                (52, 13): current_date  # 52行日期
            }
            
            # 如果有图纸变更信息，添加到PSW中
            if drawing_change:
                # 这里可以根据PSW模板的实际布局调整位置
                # 示例：在适当位置添加图纸变更信息
                fields_to_fill[(47, 4)] = f"图纸变更版本: {drawing_change.drawing_change_version}"
                fields_to_fill[(48, 4)] = f"CR号: {drawing_change.cr_number}"
                fields_to_fill[(49, 4)] = f"变更次数: {drawing_change.change_count}"
                fields_to_fill[(50, 4)] = f"变更日期: {drawing_change.change_date.strftime('%Y-%m-%d')}"

            # 填充项目信息字段
            if project:
                project_fields = {
                    (14, 3): project.supplier_name if project.supplier_name else "N/A",  # 供方名称
                    (14, 8): project.supplier_code if project.supplier_code else "N/A",  # 供方代码
                    (14, 13): project.customer_name if project.customer_name else "N/A",  # 客户名称
                    (15, 3): project.address if project.address else "N/A",  # 街道地址
                    (15, 14): project.customer_purchase if project.customer_purchase else "N/A",  # 客户采购
                    # 质量工程师信息
                    (53, 3): project.quality_engineer if project.quality_engineer else "N/A",  # 印刷体姓名（质量工程师姓名）
                    (54, 2): "质量工程师",  # 职务
                    (53, 8): project.phone if project.phone else "N/A",  # 电话号码
                    (54, 8): project.email if project.email else "N/A",  # Email
                }
                fields_to_fill.update(project_fields)

                # 处理签名图片
                if project.quality_engineer_signature:
                    try:
                        import io
                        from PIL import Image

                        # 读取签名图片
                        img_data = project.quality_engineer_signature
                        img = Image.open(io.BytesIO(img_data))

                        # 调整图片大小为固定尺寸（高度0.82厘米，宽度1.46厘米）
                        # 转换为像素（假设96 DPI）
                        width_px = int(1.46 * 96 / 2.54)  # 1.46厘米转换为像素
                        height_px = int(0.82 * 96 / 2.54)  # 0.82厘米转换为像素
                        resized_img = img.resize((width_px, height_px), Image.LANCZOS)

                        # 保存调整后的图片到临时文件
                        temp_img_path = os.path.join(tempfile.gettempdir(), f"signature_{part_id}.png")
                        resized_img.save(temp_img_path, format='PNG')

                        # 插入图片到Excel（授权的供方代表签名位置）
                        # 签名位置在第53行F列，往上一格
                        # 使用精确的尺寸：宽度1.46厘米，高度0.82厘米
                        ws.Shapes.AddPicture(
                            temp_img_path,
                            False,
                            True,
                            ws.Cells(53, 6).Left + 5,   # F列，向右偏移5像素
                            ws.Cells(53, 6).Top + 2,     # 53行，向下偏移2像素
                            width_px,                    # 使用精确的宽度
                            height_px                    # 使用精确的高度
                        )

                        # 删除临时文件
                        os.unlink(temp_img_path)
                        print("签名图片添加成功")
                    except Exception as e:
                        print(f"添加签名图片失败: {str(e)}")

            # 填充字段
            for (row, col), value in fields_to_fill.items():
                try:
                    # 直接设置单元格值，Excel会自动保留格式
                    ws.Cells(row, col).Value = value
                except Exception as e:
                    print(f"填充单元格({row},{col})失败: {str(e)}")

            # 保存并关闭工作簿
            wb.Save()
            wb.Close()
        finally:
            # 退出Excel应用程序
            excel.Quit()

        # 读取临时文件到内存
        with open(temp_file_path, 'rb') as f:
            output_data = f.read()

        # 删除临时文件
        os.unlink(temp_file_path)

        # 生成输出文件
        output_file = BytesIO(output_data)
        output_file.seek(0)

        # Send the file
        return send_file(
            output_file,
            as_attachment=True,
            download_name=f"PSW_{part.part_name}_{part.part_number}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        print(f"PSW generation error: {e}")
        import traceback
        traceback.print_exc()
        return APIResponse.internal_server_error(
            message=f"Generate PSW failed: {str(e)}",
            errors={"error": "generate PSW failed"}
        )


@api_bp.route("/drawing-changes", methods=["POST"])
def create_drawing_change():
    """Create a new drawing change"""
    try:
        data = request.get_json()
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        # Required fields
        project_id = data.get("project_id")
        part_id = data.get("part_id")
        part_name = data.get("part_name")
        part_number = data.get("part_number")
        drawing_number = data.get("drawing_number")
        engineering_change_level = data.get("engineering_change_level")
        drawing_change_version = data.get("drawing_change_version")
        cr_number = data.get("cr_number")
        change_count = data.get("change_count")
        change_date = data.get("change_date")

        if not all([project_id, part_id, part_name, part_number, drawing_number, engineering_change_level, drawing_change_version, cr_number, change_count, change_date]):
            return APIResponse.bad_request(
                message="all fields are required"
            )

        # Create new drawing change
        from datetime import datetime
        drawing_change = DrawingChange(
            project_id=project_id,
            part_id=part_id,
            part_name=part_name,
            part_number=part_number,
            drawing_number=drawing_number,
            engineering_change_level=engineering_change_level,
            drawing_change_version=drawing_change_version,
            cr_number=cr_number,
            change_count=change_count,
            change_date=datetime.strptime(change_date, "%Y-%m-%d").date()
        )

        db.add(drawing_change)
        db.commit()

        return APIResponse.created(
            data={
                "id": drawing_change.id,
                "part_name": drawing_change.part_name,
                "part_number": drawing_change.part_number,
                "drawing_number": drawing_change.drawing_number,
                "engineering_change_level": drawing_change.engineering_change_level,
                "drawing_change_version": drawing_change.drawing_change_version,
                "cr_number": drawing_change.cr_number,
                "change_count": drawing_change.change_count,
                "change_date": drawing_change.change_date.isoformat(),
                "created_at": drawing_change.created_at.isoformat() if drawing_change.created_at else None,
            }
        )
    except Exception as e:
        db.rollback()
        return APIResponse.internal_server_error(
            message=f"Create drawing change failed: {str(e)}",
            errors={"error": "create drawing change failed"}
        )


@api_bp.route("/drawing-changes", methods=["GET"])
def list_drawing_changes():
    """List all drawing changes"""
    try:
        drawing_changes = DrawingChange.query.order_by(DrawingChange.created_at.desc()).all()
        
        def ser(dc):
            return {
                "id": dc.id,
                "part_name": dc.part_name,
                "part_number": dc.part_number,
                "drawing_number": dc.drawing_number,
                "engineering_change_level": dc.engineering_change_level,
                "drawing_change_version": dc.drawing_change_version,
                "cr_number": dc.cr_number,
                "change_count": dc.change_count,
                "change_date": dc.change_date.isoformat(),
                "created_at": dc.created_at.isoformat() if dc.created_at else None,
            }

        return APIResponse.success(
            data=[ser(dc) for dc in drawing_changes]
        )
    except Exception as e:
        return APIResponse.internal_server_error(
            message=f"List drawing changes failed: {str(e)}",
            errors={"error": "list drawing changes failed"}
        )


@api_bp.route("/ods/preview/<int:part_id>", methods=["GET"])
def get_ods_preview(part_id):
    """获取ODS预览数据"""
    try:
        # 从数据库获取尺寸数据
        from bom_system.dimensions.models import Dimension
        from bom_system.dimensions.services import DimensionService
        from bom_system.models import db
        
        # 首先获取零件信息
        from bom_system.models import BomTable
        part = BomTable.query.get(part_id)
        
        if part:
            # 获取零件信息，使用零件编号查询尺寸数据
            part_number = part.part_number
            project_id = part.project_id
            
            # 打印调试信息
            print(f"查询尺寸数据: part_id={part_id}, part_number={part_number}, project_id={project_id}")
            
            # 使用DimensionService查询尺寸数据
            service = DimensionService(db.session)
            
            # 尝试通过零件编号查询尺寸数据
            dimension_records = service.get_dimensions_by_part_number(project_id, part_number)
            print(f"通过零件编号{part_number}查询到 {len(dimension_records)} 条记录")
            
            # 如果没找到，尝试通过part_id查询
            if not dimension_records:
                part_id_str = str(part_id)
                dimension_records = service.get_dimensions_by_part(part_id_str)
                print(f"通过part_id={part_id_str}查询到 {len(dimension_records)} 条记录")
            
            # 只获取与当前零件相关的图片尺寸数据
            # 查找part_id为空（未关联到特定零件）的图片尺寸
            from bom_system.dimensions.models import Dimension
            from sqlalchemy import or_, and_
            # 查找与当前零件相关的图片尺寸：要么part_id匹配，要么part_id为空（通用图片尺寸）
            related_image_dimensions = db.session.query(Dimension).filter(
                or_(
                    # 与当前零件直接关联的图片尺寸
                    and_(
                        Dimension.dimension_type.in_(['image', 'image_dimension']),
                        Dimension.part_id == part_number
                    ),
                    # 未关联到特定零件的通用图片尺寸
                    and_(
                        Dimension.dimension_type.in_(['image', 'image_dimension']),
                        Dimension.part_id == ""
                    )
                )
            ).all()
            print(f"找到 {len(related_image_dimensions)} 条与当前零件相关的图片尺寸记录")
            
            # 合并尺寸数据，避免重复
            all_dimensions = dimension_records.copy()
            existing_ids = {dim.id for dim in dimension_records}
            for img_dim in related_image_dimensions:
                if img_dim.id not in existing_ids:
                    all_dimensions.append(img_dim)
            
            # 如果仍然没有数据，尝试获取项目所有尺寸
            if not all_dimensions:
                all_dimensions = service.get_dimensions_by_project(project_id)
                print(f"通过项目ID{project_id}查询到 {len(all_dimensions)} 条记录")
            
            dimension_records = all_dimensions
            
            # 关闭数据库会话
            db.session.close()
        else:
            # 如果找不到零件，返回空列表
            print(f"未找到零件: part_id={part_id}")
            dimension_records = []
        
        # 转换数据格式
        dimensions = []
        for dim in dimension_records:
            # 处理特征描述
            technical_note = ""
            
            # 检查是否为图片尺寸类型
            if dim.dimension_type in ['image', 'image_dimension']:
                # 图片尺寸显示为"图片尺寸"
                technical_note = "图片尺寸"
            else:
                # 使用尺寸格式化工具
                from bom_system.dimensions.formatters import format_dimension
                technical_note = format_dimension(dim)
            
            dimensions.append({
                "sequence_no": dim.id,
                "technical_note": technical_note,
                "measurement_method": "卡尺/游标卡尺",
                "characteristic_code": dim.characteristic or "",
                "frequency": "首/巡/末检",
                "image_url": dim.image_url or ""
            })
        
        # 如果没有尺寸数据，使用默认数据
        if not dimensions:
            dimensions = [
                {
                    "sequence_no": 1,
                    "technical_note": "尺寸1: 100±0.1",
                    "measurement_method": "卡尺",
                    "characteristic_code": "CR",
                    "frequency": "首/巡/末检"
                },
                {
                    "sequence_no": 2,
                    "technical_note": "尺寸2: 50±0.05",
                    "measurement_method": "游标卡尺",
                    "characteristic_code": "MA",
                    "frequency": "首/巡/末检"
                }
            ]
        
        return APIResponse.success(
            data={
                "dimensions": dimensions,
                "part_id": part_id
            }
        )
    except Exception as e:
        print(f"获取ODS预览数据失败: {str(e)}")
        return APIResponse.internal_server_error(
            message=f"Get ODS preview failed: {str(e)}",
            errors={"error": "get ods preview failed"}
        )

@api_bp.route("/ods/generate", methods=["POST"])
def generate_ods():
    """生成ODS检验指导书"""
    try:
        print("开始生成ODS")
        data = request.get_json()
        print(f"获取到数据: {data}")
        if not data:
            return APIResponse.bad_request(
                message="request body required"
            )

        part_id = data.get("part_id")
        header_data = data.get("header_data", {})
        print(f"part_id: {part_id}, header_data: {header_data}")

        if not part_id:
            return APIResponse.bad_request(
                message="part_id is required"
            )

        # 确保part_id是整数
        try:
            part_id = int(part_id)
        except ValueError:
            return APIResponse.bad_request(
                message="part_id must be an integer"
            )

        # 获取零件信息
        part = BomTable.query.get(part_id)
        print(f"获取到零件信息: {part}")
        if not part:
            return APIResponse.not_found(
                message="part not found"
            )

        # 获取项目信息
        project = Project.query.get(part.project_id)
        print(f"获取到项目信息: {project}")
        if not project:
            return APIResponse.not_found(
                message="project not found"
            )

        # 构建项目信息
        project_info = {
            "id": project.id,
            "name": project.name,
            "customer_name": project.customer_name,
            "description": project.description,
        }
        print(f"构建项目信息: {project_info}")

        # 构建零件信息
        part_info = {
            "part_name": part.part_name,
            "part_number": part.part_number,
            "drawing_2d": part.drawing_2d,
            "drawing_3d": part.drawing_3d,
            "original_material": part.original_material,
            "final_material_cn": part.final_material_cn,
        }
        print(f"构建零件信息: {part_info}")

        # 从数据库获取尺寸数据
        from bom_system.dimensions.models import Dimension
        from bom_system.dimensions.services import DimensionService
        
        # 获取零件信息，使用零件编号查询尺寸数据
        part_number = part.part_number
        project_id = part.project_id
        
        # 打印调试信息
        print(f"查询尺寸数据: part_id={part_id}, part_number={part_number}, project_id={project_id}")
        
        # 使用DimensionService查询尺寸数据
        from bom_system.models import db
        service = DimensionService(db.session)
        
        # 尝试通过零件编号查询尺寸数据
        dimension_records = service.get_dimensions_by_part_number(project_id, part_number)
        print(f"通过零件编号{part_number}查询到 {len(dimension_records)} 条记录")
        
        # 如果没找到，尝试通过part_id查询
        if not dimension_records:
            part_id_str = str(part_id)
            dimension_records = service.get_dimensions_by_part(part_id_str)
            print(f"通过part_id={part_id_str}查询到 {len(dimension_records)} 条记录")
        
        # 只获取与当前零件相关的图片尺寸数据
        # 查找part_id为空（未关联到特定零件）的图片尺寸
        from bom_system.dimensions.models import Dimension
        from sqlalchemy import or_, and_
        # 查找与当前零件相关的图片尺寸：要么part_id匹配，要么part_id为空（通用图片尺寸）
        related_image_dimensions = db.session.query(Dimension).filter(
            or_(
                # 与当前零件直接关联的图片尺寸
                and_(
                    Dimension.dimension_type.in_(['image', 'image_dimension']),
                    Dimension.part_id == part_number
                ),
                # 未关联到特定零件的通用图片尺寸
                and_(
                    Dimension.dimension_type.in_(['image', 'image_dimension']),
                    Dimension.part_id == ""
                )
            )
        ).all()
        print(f"找到 {len(related_image_dimensions)} 条与当前零件相关的图片尺寸记录")
        
        # 合并尺寸数据，避免重复
        all_dimensions = dimension_records.copy()
        existing_ids = {dim.id for dim in dimension_records}
        for img_dim in related_image_dimensions:
            if img_dim.id not in existing_ids:
                all_dimensions.append(img_dim)
        
        # 如果仍然没有数据，尝试获取项目所有尺寸
        if not all_dimensions:
            all_dimensions = service.get_dimensions_by_project(project_id)
            print(f"通过项目ID{project_id}查询到 {len(all_dimensions)} 条记录")
        
        dimension_records = all_dimensions
        
        # 关闭数据库会话
        db.session.close()
        
        # 转换数据格式
        dimensions = []
        print(f"dimension_records类型: {type(dimension_records)}")
        print(f"dimension_records长度: {len(dimension_records)}")
        for i, dim in enumerate(dimension_records):
            print(f"dimension_records[{i}]类型: {type(dim)}")
            # 处理特征描述
            drawing = ""
            
            # 检查是否为图片尺寸类型
            if hasattr(dim, 'dimension_type'):
                if dim.dimension_type in ['image', 'image_dimension']:
                    # 图片尺寸显示为"图片尺寸"
                    drawing = "图片尺寸"
                else:
                    # 使用尺寸格式化工具
                    from bom_system.dimensions.formatters import format_dimension
                    drawing = format_dimension(dim)
            
            dim_dict = {
                "drawing": drawing,
                "method": "卡尺/游标卡尺",
                "special": dim.characteristic or "" if hasattr(dim, 'characteristic') else "",
                "frequency": "首/巡/末检",
                "imageUrl": dim.image_url or "" if hasattr(dim, 'image_url') else ""
            }
            print(f"添加到dimensions的元素: {dim_dict}")
            dimensions.append(dim_dict)
        
        # 如果没有尺寸数据，使用默认数据
        if not dimensions:
            dimensions = [
                {
                    "drawing": "尺寸1: 100±0.1",
                    "method": "卡尺",
                    "special": "CR",
                    "frequency": "首/巡/末检"
                },
                {
                    "drawing": "尺寸2: 50±0.05",
                    "method": "游标卡尺",
                    "special": "MA",
                    "frequency": "首/巡/末检"
                }
            ]

        # 直接创建简单的ODS文件
        import tempfile
        from openpyxl import Workbook
        
        # 创建新工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ODS数据"
        
        # 添加标题
        sheet["A1"] = f"项目：{project_info.get('name', '')}"
        sheet["A2"] = f"客户：{project_info.get('customer_name', '')}"
        
        # 添加表头
        headers = ["序号", "特征描述", "检测方法", "特性号码", "频率"]
        for col, header in enumerate(headers, 1):
            sheet.cell(4, col).value = header
        
        # 添加数据
        for row, dim in enumerate(dimensions, 5):
            sheet.cell(row, 1).value = row - 4  # 序号
            sheet.cell(row, 2).value = dim.get("drawing", "")  # 特征描述
            sheet.cell(row, 3).value = dim.get("method", "")  # 检测方法
            sheet.cell(row, 4).value = dim.get("special", "")  # 特性号码
            sheet.cell(row, 5).value = dim.get("frequency", "")  # 频率
        
        # 保存到临时文件
        output_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        output_file.close()
        workbook.save(output_file.name)

        # 发送文件
        return send_file(
            output_file.name,
            as_attachment=True,
            download_name=f"ODS_{part.part_name}_{part.part_number}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        print(f"ODS generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        # 打印完整的错误堆栈
        import sys
        exc_type, exc_value, exc_traceback = sys.exc_info()
        error_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
        error_message = ''.join(error_lines)
        print(f"完整错误信息: {error_message}")
        return APIResponse.internal_server_error(
            message=f"Generate ODS failed: {str(e)}",
            errors={"error": "generate ODS failed"}
        )


@api_bp.route("/process-capability/generate", methods=["POST"])
def generate_process_capability():
    """生成初始过程能力分析报告"""
    print("=== 开始处理初始过程能力分析报告请求 ===")
    try:
        data = request.get_json()
        print(f"收到的请求数据: {data}")
        if not data:
            print("错误: 缺少请求体")
            return APIResponse.bad_request(
                message="request body required"
            )

        part_id = data.get("part_id")
        project_id = data.get("project_id")
        print(f"part_id: {part_id}, project_id: {project_id}")

        if not part_id and not project_id:
            print("错误: 缺少part_id或project_id")
            return APIResponse.bad_request(
                message="either part_id or project_id is required"
            )

        # 模板文件路径
        template_path = r"C:\Users\Administrator\univer\output\01-初始过程能力分析报告 Y1393847.xls"
        print(f"模板文件路径: {template_path}")
        print(f"模板文件存在: {os.path.exists(template_path)}")
        if not os.path.exists(template_path):
            print("错误: 模板文件不存在")
            return APIResponse.internal_server_error(
                message="template file not found"
            )

        # 导入ProcessCapabilityService
        from bom_system.process_capability import ProcessCapabilityService
        print("导入ProcessCapabilityService成功")
        service = ProcessCapabilityService(template_path)
        print("创建ProcessCapabilityService实例成功")

        # 生成输出目录
        output_dir = os.path.join(os.path.dirname(template_path), "process_capability")
        print(f"创建输出目录: {output_dir}")
        os.makedirs(output_dir, exist_ok=True)
        print(f"输出目录创建成功: {os.path.exists(output_dir)}")

        if part_id:
            # 为单个零件生成报告
            try:
                part_id = int(part_id)
                print(f"转换part_id为整数: {part_id}")
            except ValueError:
                print("错误: part_id不是有效的整数")
                return APIResponse.bad_request(
                    message="part_id must be an integer"
                )

            from .models import BomTable
            part = BomTable.query.get(part_id)
            if not part:
                print(f"错误: 找不到零件，part_id: {part_id}")
                return APIResponse.not_found(
                    message="part not found"
                )

            print(f"找到零件: id={part.id}, part_number={part.part_number}, part_name={part.part_name}")

            # 生成报告
            print(f"开始生成报告，零件号: {str(part.part_number)}")
            try:
                output_paths = service.generate_report(part, output_dir)
                print(f"报告生成成功: {len(output_paths)} 个文件")
                for path in output_paths:
                    print(f"生成的文件: {str(path)}")
            except Exception as e:
                print(f"生成报告时出错: {str(e)}")
                import traceback
                traceback.print_exc()
                return APIResponse.internal_server_error(
                    message=f"Generate report failed: {str(e)}",
                    errors={"error": "generate report failed"}
                )
            
            # 检查生成的文件是否存在
            if not output_paths:
                print("错误: 没有生成任何文件")
                return APIResponse.internal_server_error(
                    message="无法为该零件生成过程能力分析报告，可能是因为缺少尺寸数据",
                    errors={"error": "no files generated"}
                )
            
            for output_path in output_paths:
                if not os.path.exists(output_path):
                    print(f"文件不存在: {str(output_path)}")
                    return APIResponse.internal_server_error(
                        message=f"Generated file not found: {str(output_path)}",
                        errors={"error": "generated file not found"}
                    )
            
            # 读取文件内容
            output_path = output_paths[0]
            print(f"读取文件: {output_path}")
            try:
                with open(output_path, 'rb') as f:
                    file_content = f.read()
                print(f"文件大小: {len(file_content)} bytes")
            except Exception as e:
                print(f"读取文件时出错: {str(e)}")
                return APIResponse.internal_server_error(
                    message=f"Read file failed: {str(e)}",
                    errors={"error": "read file failed"}
                )
            
            # 创建响应对象
            from flask import make_response
            response = make_response(file_content)
            
            # 设置响应头
            import urllib.parse
            filename = os.path.basename(output_path)
            encoded_filename = urllib.parse.quote(filename)
            print(f"文件名: {filename}, 编码后: {encoded_filename}")
            
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=UTF-8''{encoded_filename}'
            response.headers['Content-Length'] = str(len(file_content))
            
            print("=== 请求处理完成 ===")
            # 发送文件
            return response
        elif project_id:
            # 为项目中的所有零件生成报告
            try:
                project_id = int(project_id)
            except ValueError:
                print("错误: project_id不是有效的整数")
                return APIResponse.bad_request(
                    message="project_id must be an integer"
                )

            # 生成报告
            try:
                generated_files = service.generate_reports_for_project(project_id, output_dir)
                print(f"为项目生成报告成功: {len(generated_files)} 个文件")
                for path in generated_files:
                    print(f"生成的文件: {str(path)}")
            except Exception as e:
                print(f"生成报告时出错: {str(e)}")
                import traceback
                traceback.print_exc()
                return APIResponse.internal_server_error(
                    message=f"Generate reports failed: {str(e)}",
                    errors={"error": "generate reports failed"}
                )

            # 检查生成的文件是否存在
            if not generated_files:
                print("错误: 项目中没有零件")
                return APIResponse.not_found(
                    message="no parts found for this project"
                )

            # 返回生成的文件列表
            return APIResponse.success(
                message=f"Generated {len(generated_files)} process capability reports",
                data={
                    "status": "success",
                    "files": [os.path.basename(f) for f in generated_files],
                    "output_dir": output_dir
                }
            )

    except Exception as e:
        print(f"Process capability generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        return APIResponse.internal_server_error(
            message=f"Generate process capability report failed: {str(e)}",
            errors={"error": "generate process capability report failed"}
        )
