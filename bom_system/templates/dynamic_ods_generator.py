#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
动态ODS生成器

基于Excel模板和数据库数据生成ODS文件
"""

import logging
import tempfile
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


class DynamicODSGenerator:
    """动态ODS生成器"""
    
    def __init__(self):
        self.logger = logger
    
    def generate_from_template(
        self,
        template_path: str,
        data: List[Dict[str, Any]],
        project_info: Dict[str, Any],
        part_info: Dict[str, Any] = None,
        config: Dict[str, Any] = None,
        output_format: str = "xlsx"
    ) -> str:
        """
        基于模板和数据生成ODS文件
        
        Args:
            template_path: 模板文件路径
            data: 零件数据列表
            project_info: 项目信息
            part_info: 零件信息
            config: 生成配置
            output_format: 输出格式 (xlsx/ods)
            
        Returns:
            生成文件的路径
        """
        try:
            self.logger.info(f"开始生成ODS文件，模板: {template_path}")
            
            # 加载模板
            workbook = load_workbook(template_path)
            
            # 获取主工作表
            main_sheet = self._find_main_sheet(workbook)
            
            # 查找数据插入位置
            data_start_row = self._find_data_start_row(main_sheet)
            
            # 填充项目信息和零件信息
            self._fill_project_info(main_sheet, project_info, part_info)
            
            # 填充零件数据
            self._fill_parts_data(main_sheet, data, data_start_row, config)
            
            # 创建临时输出文件
            output_file = tempfile.NamedTemporaryFile(
                suffix=f".{output_format}",
                delete=False
            )
            output_file.close()
            
            # 保存文件
            workbook.save(output_file.name)
            
            self.logger.info(f"ODS文件生成成功: {output_file.name}")
            
            return output_file.name
            
        except Exception as e:
            self.logger.error(f"生成ODS文件失败: {str(e)}")
            raise
    
    def _find_main_sheet(self, workbook: Workbook):
        """查找主工作表"""
        # 优先查找包含"ODS"或"主表"的工作表
        for sheet_name in workbook.sheetnames:
            if any(keyword in sheet_name.upper() for keyword in ["ODS", "主表", "MAIN", "数据"]):
                return workbook[sheet_name]
        
        # 如果没找到，使用第一个工作表
        return workbook.active
    
    def _find_data_start_row(self, sheet) -> int:
        """查找数据开始行"""
        # 简单策略：查找包含"序号"、"零件编号"等关键字的行作为表头
        for row in range(1, min(20, sheet.max_row + 1)):  # 只检查前20行
            for col in range(1, min(10, sheet.max_column + 1)):  # 只检查前10列
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in cell_value for keyword in ["序号", "零件编号", "零件名称", "图号"]):
                        return row + 1  # 返回数据开始行（表头下一行）
        
        # 如果没找到表头，假设从第5行开始
        return 5
    
    def _fill_project_info(self, sheet, project_info: Dict[str, Any], part_info: Dict[str, Any] = None):
        """填充项目信息和零件信息"""
        try:
            # 查找项目信息填充位置
            info_mappings = {
                "项目名称": project_info.get("name", ""),
                "项目编号": str(project_info.get("id", "")),
                "项目描述": project_info.get("description", ""),
                "状态": project_info.get("status", ""),
                "客户名称": project_info.get("customer_name", ""),
                "客户": project_info.get("customer_name", "")
            }
            
            # 在前几行查找对应的标签并填充
            for row in range(1, min(10, sheet.max_row + 1)):
                for col in range(1, min(10, sheet.max_column + 1)):
                    cell_value = sheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        for label, value in info_mappings.items():
                            if label in cell_value:
                                # 在下一列或下一行填充值
                                if col < sheet.max_column:
                                    sheet.cell(row, col + 1).value = value
                                break
            
            # 特殊处理WZ1D模板的固定位置
            # 项目名称：第2行第6列
            if project_info.get("name"):
                sheet.cell(2, 6).value = project_info.get("name")
            # 客户名称：第2行第8列
            if project_info.get("customer_name"):
                sheet.cell(2, 8).value = project_info.get("customer_name")
            
            # 填充零件信息
            if part_info:
                # 产品名称：第2行第10列
                if part_info.get("part_name"):
                    sheet.cell(2, 10).value = part_info.get("part_name")
                # 产品编号：第3行第10列
                if part_info.get("part_number"):
                    sheet.cell(3, 10).value = part_info.get("part_number")
                # 图纸日期/版本：第3行第8列
                if part_info.get("drawing_2d"):
                    sheet.cell(3, 8).value = part_info.get("drawing_2d")
            
        except Exception as e:
            self.logger.warning(f"填充项目信息失败: {str(e)}")
    
    def _fill_parts_data(self, sheet, data: List[Dict[str, Any]], start_row: int, config: Dict[str, Any]):
        """填充零件数据"""
        try:
            # 检查是否是WZ1D模板
            is_wz1d_template = False
            if sheet.max_row >= 5:
                header_cell = sheet.cell(5, 1).value
                if header_cell and "序号" in str(header_cell):
                    is_wz1d_template = True
            
            if is_wz1d_template:
                # WZ1D模板的特殊处理
                # 清除从第11行开始的数据，保留模板格式
                self._clear_existing_data(sheet, 11)
                
                # 填充尺寸检验数据
                for row_idx, dimension in enumerate(data, start=11):  # 从第11行开始填充尺寸检验数据
                    # 序号
                    sheet.cell(row_idx, 1).value = row_idx - 10  # 从1开始编号
                    # 关键控制事项（尺寸描述）
                    sheet.cell(row_idx, 2).value = dimension.get("drawing", "")
                    # 检查方法
                    sheet.cell(row_idx, 3).value = dimension.get("method", "卡尺/游标卡尺")
                    # 特性
                    sheet.cell(row_idx, 4).value = dimension.get("special", "")
                    # 频率
                    sheet.cell(row_idx, 5).value = dimension.get("frequency", "首/巡/末检")
                    
                    # 处理图片尺寸
                    if dimension.get("imageUrl"):
                        try:
                            from openpyxl.drawing.image import Image
                            from io import BytesIO
                            import requests
                            
                            # 下载图片
                            image_url = dimension.get("imageUrl")
                            # 如果是相对路径，添加完整URL
                            if image_url.startswith('/'):
                                image_url = f"http://localhost:5000{image_url}"
                            
                            response = requests.get(image_url)
                            if response.status_code == 200:
                                # 创建图片对象
                                img = Image(BytesIO(response.content))
                                
                                # 调整图片大小
                                img.width = 100  # 设置图片宽度
                                img.height = 80   # 设置图片高度
                                
                                # 将图片插入到第6列
                                sheet.add_image(img, f"F{row_idx}")
                                self.logger.info(f"成功插入图片到行 {row_idx}")
                        except Exception as e:
                            self.logger.warning(f"插入图片失败: {str(e)}")
                
                self.logger.info(f"成功填充 {len(data)} 行尺寸检验数据到WZ1D模板")
            else:
                # 通用模板处理
                # 定义列映射
                column_mappings = {
                    "A": "sequence",           # 序号
                    "B": "part_number",        # 零件编号
                    "C": "part_name",          # 零件名称
                    "D": "drawing_2d",         # 2D图号
                    "E": "drawing_3d",         # 3D图号
                    "F": "original_material",  # 原材料
                    "G": "final_material_cn",  # 终审材料
                    "H": "part_category",      # 零件分类
                    "I": "net_weight_kg",      # 净重
                }
                
                # 清除现有数据行
                self._clear_existing_data(sheet, start_row)
                
                # 填充新数据
                for row_idx, part in enumerate(data, start=start_row):
                    for col_letter, field_name in column_mappings.items():
                        value = part.get(field_name, "")
                        
                        # 特殊处理
                        if field_name == "sequence" and not value:
                            value = row_idx - start_row + 1  # 自动编号
                        elif field_name == "net_weight_kg" and value:
                            try:
                                value = float(value)
                            except (ValueError, TypeError):
                                value = ""
                        
                        sheet[f"{col_letter}{row_idx}"].value = value
                        
                self.logger.info(f"成功填充 {len(data)} 行零件数据")
            
        except Exception as e:
            self.logger.error(f"填充零件数据失败: {str(e)}")
            raise
    
    def _clear_existing_data(self, sheet, start_row: int):
        """清除现有数据行，但保留格式"""
        try:
            # 只清除数据，不删除行，保留格式
            max_row = sheet.max_row
            max_col = sheet.max_column
            if max_row >= start_row:
                for row in range(start_row, max_row + 1):
                    for col in range(1, max_col + 1):
                        sheet.cell(row, col).value = None
        except Exception as e:
            self.logger.warning(f"清除现有数据失败: {str(e)}")
    
    def _apply_formatting(self, sheet, start_row: int, data_count: int):
        """应用格式化"""
        try:
            # 这里可以添加格式化逻辑
            # 例如：边框、字体、对齐等
            pass
        except Exception as e:
            self.logger.warning(f"应用格式化失败: {str(e)}")
    
    def _add_images(self, sheet, data: List[Dict[str, Any]], config: Dict[str, Any]):
        """添加图片（如果配置启用）"""
        try:
            if not config or not config.get("includeImages", False):
                return
                
            # 这里可以添加图片插入逻辑
            # 由于复杂性，暂时跳过
            self.logger.info("图片插入功能暂未实现")
            
        except Exception as e:
            self.logger.warning(f"添加图片失败: {str(e)}")


# 辅助函数
def create_simple_ods_from_data(data: List[Dict[str, Any]], project_info: Dict[str, Any]) -> str:
    """
    创建简单的ODS文件（不使用模板）
    
    Args:
        data: 零件数据
        project_info: 项目信息
        
    Returns:
        生成文件路径
    """
    try:
        # 创建新工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ODS数据"
        
        # 添加标题
        sheet["A1"] = f"项目：{project_info.get('name', '')}"
        sheet["A2"] = f"客户：{project_info.get('customer_name', '')}"
        sheet["A3"] = f"生成时间：{project_info.get('generated_time', '')}"
        
        # 添加表头
        headers = [
            "序号", "零件编号", "零件名称", "2D图号", "3D图号", 
            "原材料", "终审材料", "零件分类", "净重(kg)"
        ]
        
        for col, header in enumerate(headers, 1):
            sheet.cell(4, col).value = header
        
        # 添加数据
        for row, part in enumerate(data, 5):
            sheet.cell(row, 1).value = row - 4  # 序号
            sheet.cell(row, 2).value = part.get("part_number", "")
            sheet.cell(row, 3).value = part.get("part_name", "")
            sheet.cell(row, 4).value = part.get("drawing_2d", "")
            sheet.cell(row, 5).value = part.get("drawing_3d", "")
            sheet.cell(row, 6).value = part.get("original_material", "")
            sheet.cell(row, 7).value = part.get("final_material_cn", "")
            sheet.cell(row, 8).value = part.get("part_category", "")
            sheet.cell(row, 9).value = part.get("net_weight_kg", "")
        
        # 保存到临时文件
        output_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        output_file.close()
        workbook.save(output_file.name)
        
        return output_file.name
        
    except Exception as e:
        logger.error(f"创建简单ODS文件失败: {str(e)}")
        raise
