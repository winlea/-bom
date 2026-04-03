# -*- coding: utf-8 -*-
from __future__ import annotations

import glob
import io
import os
import re
import shutil
import subprocess
import tempfile
import urllib.request
import uuid
from copy import copy
from typing import Any, Dict, List, Optional, Tuple

from flask import Blueprint, Response, current_app, jsonify, redirect, request
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from bom_system.database.session import get_db_session


# 可选：文本渲染为 PNG（DRAW_DIM 无图时回退）
try:
    from PIL import Image as PILImage
    from PIL import ImageDraw, ImageFont
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

from bom_system.dimensions.services import DimensionService

# 业务服务
from bom_system.models import db  # noqa: F401

templates_bp = Blueprint("templates_bp", __name__, url_prefix="/api/templates")


# ---------------- 通用工具 ----------------
def _project_root() -> str:
    return os.path.abspath(current_app.root_path)


def _ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def _safe_template_path(rel_path: str) -> str:
    proj_root = _project_root()
    templates_dir = os.path.join(proj_root, "templates")
    rel_path = rel_path.replace("\\", "/")
    if rel_path.startswith("/"):
        rel_path = rel_path[1:]
    if not (rel_path.startswith("templates/") or rel_path == "templates"):
        raise ValueError("非法路径：仅允许访问 templates 目录")
    abs_path = os.path.abspath(os.path.join(proj_root, rel_path))
    if not abs_path.startswith(os.path.abspath(templates_dir)):
        raise ValueError("非法路径：越权访问")
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"文件不存在: {abs_path}")
    return abs_path


def _build_merge_map(ws: Worksheet) -> Dict[Tuple[int, int], Dict[str, int]]:
    merge_map: Dict[Tuple[int, int], Dict[str, int]] = {}
    for cr in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = cr.min_row, cr.min_col, cr.max_row, cr.max_col
        rowspan = max_r - min_r + 1
        colspan = max_c - min_c + 1
        merge_map[(min_r, min_c)] = {"rowspan": rowspan, "colspan": colspan}
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) == (min_r, min_c):
                    continue
                merge_map[(r, c)] = {"skip": 1}
    return merge_map


def _excel_col_width_to_px(width: Optional[float]) -> int:
    if width is None:
        return 80
    try:
        px = int(round((width + 0.83) * 7))
        return max(40, min(px, 800))
    except Exception:
        return 80


def _points_to_px(points: Optional[float]) -> int:
    if not points:
        return 24
    return int(round(points * 96 / 72))


def _color_to_css(c: Optional[Color]) -> Optional[str]:
    if not c:
        return None
    rgb = getattr(c, "rgb", None)
    if rgb and isinstance(rgb, str):
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    return None


def _fill_bg_to_css(fill: Optional[PatternFill]) -> Optional[str]:
    if not fill:
        return None
    if getattr(fill, "patternType", None) not in ("solid", "SOLID"):
        return None
    fg = getattr(fill, "fgColor", None)
    return _color_to_css(fg)


def _alignment_to_css(align: Optional[Alignment]) -> Dict[str, str]:
    css: Dict[str, str] = {}
    if not align:
        return css
    h = align.horizontal
    if h in ("left", "center", "right", "justify"):
        css["text-align"] = "center" if h == "center" else h
    v = align.vertical
    if v in ("top", "center", "bottom"):
        css["vertical-align"] = "middle" if v == "center" else v
    if align.wrapText:
        css["white-space"] = "pre-wrap"
        css["word-break"] = "break-word"
    return css


def _font_to_css(font: Optional[Font]) -> Dict[str, str]:
    css: Dict[str, str] = {}
    if not font:
        return css
    if font.bold:
        css["font-weight"] = "700"
    if font.italic:
        css["font-style"] = "italic"
    if font.sz:
        css["font-size"] = f"{int(round(float(font.sz) * 96 / 72))}px"
    color = _color_to_css(font.color)
    if color:
        css["color"] = color
    return css


def _style_dict_to_str(style: Dict[str, str]) -> str:
    if not style:
        return ""
    return "; ".join([f"{k}: {v}" for k, v in style.items()])


# ---------------- 纯 Python 预览（辅助） ----------------
def _sheet_to_html(ws: Worksheet) -> str:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    merge_map = _build_merge_map(ws)

    html = io.StringIO()
    html.write(f'<div class="sheet-block"><h3 class="sheet-title">{ws.title}</h3>')
    html.write('<table class="sheet" cellspacing="0" cellpadding="0">')
    html.write("<colgroup>")
    for c in range(1, max_col + 1):
        letter = ws.cell(row=1, column=c).column_letter
        width_px = _excel_col_width_to_px(ws.column_dimensions.get(letter).width if letter in ws.column_dimensions else None)
        html.write(f'<col style="width: {width_px}px" />')
    html.write("</colgroup>")

    for r in range(1, max_row + 1):
        row_dim = ws.row_dimensions.get(r)
        row_height_px = _points_to_px(getattr(row_dim, "height", None))
        html.write(f'<tr style="height: {row_height_px}px">')
        for c in range(1, max_col + 1):
            info = merge_map.get((r, c))
            if info and "skip" in info:
                continue

            cell: Cell = ws.cell(row=r, column=c)
            value = cell.value
            text = "" if value is None else str(value)

            td_style: Dict[str, str] = {"border": "1px solid #e2e8f0", "padding": "0"}
            bg = _fill_bg_to_css(cell.fill)
            if bg:
                td_style["background-color"] = bg
            td_style.update(_alignment_to_css(cell.alignment))
            if "vertical-align" not in td_style:
                td_style["vertical-align"] = "middle"

            inner_style: Dict[str, str] = {
                "padding": "6px 8px",
                "line-height": "1.4",
                "white-space": "pre-wrap",
                "word-break": "break-word",
                "min-height": "20px",
                "overflow": "hidden",
                "text-overflow": "clip",
            }
            inner_style.update(_font_to_css(cell.font))

            attrs = []
            if info:
                if "rowspan" in info and info["rowspan"] > 1:
                    attrs.append(f'rowspan="{info["rowspan"]}"')
                if "colspan" in info and info["colspan"] > 1:
                    attrs.append(f'colspan="{info["colspan"]}"')
            attr_str = " " + " ".join(attrs) if attrs else ""

            html.write(f'<td{attr_str} style="{_style_dict_to_str(td_style)}">')
            html.write(f'<div class="cell" style="{_style_dict_to_str(inner_style)}">{text}</div>')
            html.write("</td>")
        html.write("</tr>")
    html.write("</table></div>")
    return html.getvalue()


def _wrap_html(body: str) -> str:
    return f"""<!doctype html>
<html lang="zh-cn">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>模板预览</title>
<style>
  html, body{{height:100%;}}
  body{{font-family: system-ui, -apple-system, "Segoe UI", Roboto, Arial, "PingFang SC","Hiragino Sans GB","Microsoft YaHei",sans-serif; background:#f8fafc; color:#0f172a; margin:0; padding:16px;}}
  .container{{max-width:1400px; margin:0 auto;}}
  .sheet-block{{background:#fff; border:1px solid #e2e8f0; border-radius:8px; padding:12px; margin:16px 0; box-shadow:0 1px 2px rgba(0,0,0,.04); overflow:auto;}}
  .sheet-title{{margin:4px 0 12px; font-size:14px; color:#475569}}
  table.sheet{{width:100%; border-collapse:collapse; table-layout:fixed; background:#fff;}}
  table.sheet td{{box-sizing:border-box}}
  .cell{{box-sizing:border-box}}
</style>
</head>
<body>
<div class="container">
{body}
</div>
</body>
</html>"""


@templates_bp.get("/html")
def get_template_as_html():
    rel_path = request.args.get("path", "templates/WZ1D_standard_template.xlsx")
    try:
        abs_path = _safe_template_path(rel_path)
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 400

    try:
        wb = load_workbook(abs_path, data_only=True)
    except Exception as e:
        return jsonify({"success": False, "message": f"读取模板失败: {e}"}), 500

    parts = [_sheet_to_html(ws) for ws in wb.worksheets]
    html = _wrap_html("\n".join(parts))
    return Response(html, mimetype="text/html")


# ---------------- LibreOffice 预览 ----------------
def _soffice_path() -> str:
    soffice = os.environ.get("LIBREOFFICE_PATH") or "C:\\Program Files\\LibreOffice\\program\\soffice.exe"
    if not os.path.exists(soffice):
        alt = "C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe"
        if os.path.exists(alt):
            soffice = alt
    if not os.path.exists(soffice):
        soffice = "soffice"
    return soffice


def _run_soffice_convert(in_file: str, out_dir: str, fmt: str = "html") -> Tuple[int, str, str]:
    _ensure_dir(out_dir)
    soffice = _soffice_path()
    # 避免首次启动卡顿，隔离用户配置
    temp_profile = os.path.join(tempfile.gettempdir(), f"lo_profile_{uuid.uuid4().hex[:8]}")
    os.makedirs(temp_profile, exist_ok=True)
    env = os.environ.copy()
    env["UserInstallation"] = f"file:///{temp_profile.replace(os.sep, '/')}"
    cmd = [
        soffice, "--headless", "--nologo", "--nofirststartwizard", "--norestore",
        "--convert-to", fmt, "--outdir", out_dir, in_file,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=300, env=env)
    try:
        shutil.rmtree(temp_profile, ignore_errors=True)
    except Exception:
        pass
    return proc.returncode, proc.stdout, proc.stderr


@templates_bp.get("/lo-preview")
def lo_preview():
    rel_path = request.args.get("path", "templates/WZ1D_standard_template.xlsx")
    try:
        abs_in = _safe_template_path(rel_path)
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 400

    root = _project_root()
    out_dir = os.path.join(root, "static", "lo-preview", uuid.uuid4().hex[:12])
    os.makedirs(out_dir, exist_ok=True)

    try:
        code, stdout, stderr = _run_soffice_convert(abs_in, out_dir, "html")
    except Exception as e:
        return jsonify({"success": False, "message": f"执行 LibreOffice 失败: {e}"}), 500

    html_files = sorted(glob.glob(os.path.join(out_dir, "*.html")))
    if not html_files:
        return jsonify({"success": False, "message": f"未生成 HTML。stdout={stdout} stderr={stderr}"}), 500

    html_name = os.path.basename(html_files[0])
    static_url = f"/static/lo-preview/{os.path.basename(os.path.dirname(html_files[0]))}/{html_name}"
    return redirect(static_url, code=302)


# ---------------- 模板模块化：inspect / fill ----------------
_PLACEHOLDER_RE = re.compile(r"\$\{([A-Za-z0-9_]+)\}")


def _scan_placeholders(wb: Workbook) -> List[str]:
    found: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str):
                    for m in _PLACEHOLDER_RE.finditer(cell.value):
                        found.add(m.group(1))
    return sorted(found)


def _find_named_row(wb: Workbook, name: str) -> Optional[Tuple[Worksheet, int, int, int]]:
    dn = wb.defined_names.get(name)
    if not dn:
        return None
    dests = list(dn.destinations)
    if not dests:
        return None
    sheet_name, ref = dests[0]
    ws = wb[sheet_name]
    # 兼容单元格/整行/区域
    m = re.match(r"(\$?[A-Z]+\$?\d+):(\$?[A-Z]+\$?\d+)", ref)
    start = end = ref
    if m:
        start = m.group(1)
        end = m.group(2)

    def parse_addr(addr: str) -> Tuple[int, int]:
        addr = addr.replace("$", "")
        mo = re.match(r"([A-Z]+)(\d+)", addr)
        assert mo
        col = 0
        for ch in mo.group(1):
            col = col * 26 + (ord(ch) - ord('A') + 1)
        row = int(mo.group(2))
        return row, col

    r1, c1 = parse_addr(start)
    r2, c2 = parse_addr(end)
    row = min(r1, r2)
    return ws, row, min(c1, c2), max(c1, c2)


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, min_col: int, max_col: int) -> None:
    for c in range(min_col, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)
    if ws.row_dimensions.get(src_row):
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _resolve_asset_to_path(url_or_path: str) -> Optional[str]:
    if not url_or_path:
        return None
    url_or_path = url_or_path.strip()
    root = _project_root()
    if url_or_path.startswith("/static/"):
        p = os.path.join(root, url_or_path.lstrip("/"))
        return p if os.path.exists(p) else None
    if os.path.isabs(url_or_path):
        return url_or_path if os.path.exists(url_or_path) else None
    cand = os.path.join(root, url_or_path)
    if os.path.exists(cand):
        return cand
    if url_or_path.startswith("http://") or url_or_path.startswith("https://"):
        tmp_dir = os.path.join(root, "static", "exports", "tmp")
        _ensure_dir(tmp_dir)
        fname = f"img_{uuid.uuid4().hex[:8]}"
        ext = os.path.splitext(url_or_path.split("?")[0])[1] or ".png"
        dst = os.path.join(tmp_dir, fname + ext)
        try:
            urllib.request.urlretrieve(url_or_path, dst)
            return dst
        except Exception:
            return None
    return None


def _render_text_image(text: str, out_path: str, max_w: int = 360, max_h: int = 120) -> Optional[str]:
    if not PIL_AVAILABLE:
        return None
    try:
        img = PILImage.new("RGBA", (max_w, max_h), (255, 255, 255, 0))
        draw = ImageDraw.Draw(img)
        font = None
        for f in ["C:/Windows/Fonts/msyh.ttc", "C:/Windows/Fonts/msyh.ttf", "C:/Windows/Fonts/arial.ttf"]:
            try:
                font = ImageFont.truetype(f, 24)
                break
            except Exception:
                continue
        if not font:
            font = ImageFont.load_default()
        # 文本居中
        bbox = draw.textbbox((0, 0), text, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        x = max(0, (max_w - tw) // 2)
        y = max(0, (max_h - th) // 2)
        draw.text((x, y), text, fill=(20, 20, 20, 255), font=font)
        img.save(out_path)
        return out_path
    except Exception:
        return None


def _format_dim_text(d: Any) -> str:
    # 维度文本：前缀(⌀/R) + 名义值 + 公差
    nominal = (d.nominal_value or "").strip() if hasattr(d, "nominal_value") else ""
    up = (d.upper_tolerance or "").strip() if hasattr(d, "upper_tolerance") else ""
    low = (d.lower_tolerance or "").strip() if hasattr(d, "lower_tolerance") else ""
    tol = (d.tolerance_value or "").strip() if hasattr(d, "tolerance_value") else ""

    prefix = ""
    dtype = (d.dimension_type or "").lower() if hasattr(d, "dimension_type") else ""
    if dtype in ("diameter", "dia", "ø", "⌀"):
        prefix = "⌀"
    elif dtype in ("radius", "r"):
        prefix = "R"

    tol_str = ""
    if up and low:
        tol_str = f" +{up}/-{low}"
    elif tol:
        tol_str = f" ±{tol}"

    return f"{prefix}{nominal}{tol_str}".strip()


@templates_bp.get("/inspect")
def inspect_template():
    rel_path = request.args.get("path", "templates/WZ1D_standard_template.xlsx")
    try:
        abs_in = _safe_template_path(rel_path)
        src_path = abs_in

        wb = load_workbook(src_path)
    except Exception as e:
        return jsonify({"success": False, "message": f"读取模板失败: {e}"}), 400

    holders = _scan_placeholders(wb)
    row_info = _find_named_row(wb, "NR_DIM_TABLE_ROW")
    row = None
    if row_info:
        ws, r, c1, c2 = row_info
        row = {
            "sheet": ws.title,
            "row": r,
            "minCol": c1,
            "maxCol": c2,
            "columns": [get_column_letter(c) for c in range(c1, c2 + 1)],
        }
    return jsonify({"success": True, "placeholders": holders, "row": row})


@templates_bp.post("/fill")
def fill_template():
    session = get_db_session()
    data = request.get_json(silent=True) or {}
    rel_path = data.get("templatePath") or request.args.get("path") or "templates/WZ1D_standard_template.xlsx"
    project_id = data.get("projectId") or request.args.get("projectId")
    single = data.get("single") or {}
    preview = str(data.get("preview") or request.args.get("preview") or "false").lower() == "true"

    if not project_id:
        return jsonify({"success": False, "message": "缺少参数 projectId"}), 400

    try:
        abs_in = _safe_template_path(rel_path)
        src_path = abs_in

        wb: Workbook = load_workbook(src_path)
    except Exception as e:
        return jsonify({"success": False, "message": f"读取模板失败: {e}"}), 400

    # 单值占位符替换（全表查找替换）
    singles_default = {
        "PROJECT_NAME": single.get("PROJECT_NAME") or "",
        "CUSTOMER_NAME": single.get("CUSTOMER_NAME") or "",
        "PRODUCT_NAME": single.get("PRODUCT_NAME") or "",
        "PRODUCT_NO": single.get("PRODUCT_NO") or "",
    }

    def _replace_single_values(workbook: Workbook, mapping: Dict[str, str]):
        for ws in workbook.worksheets:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    if isinstance(cell.value, str) and "${" in cell.value:
                        text = cell.value
                        for k, v in mapping.items():
                            text = text.replace("${" + k + "}", str(v))
                        cell.value = text

    _replace_single_values(wb, singles_default)

    # 查找样例行
    named = _find_named_row(wb, "NR_DIM_TABLE_ROW")
    if not named:
        return jsonify({"success": False, "message": "未找到命名区域 NR_DIM_TABLE_ROW（请在模板中定义样例行）"}), 400
    ws, sample_row, c1, c2 = named

    # 读取尺寸数据并分组为“组号-序号”
    svc = DimensionService(session)
    part_id = data.get("partId") or request.args.get("partId")
    if part_id:
        dims = svc.get_dimensions_by_part(part_id)
    else:
        dims = svc.get_dimensions_by_project(project_id)
    groups: Dict[int, List[Any]] = {}
    for d in dims:
        g = getattr(d, "group_no", 0) or 0
        groups.setdefault(g, []).append(d)
    # 组内排序稳定（按 id 或 characteristic）
    for g in groups:
        groups[g].sort(key=lambda x: getattr(x, "id", 0) or 0)

    # 将样例行用于第一条，其余插入
    all_items: List[Any] = []
    for g in sorted(groups.keys()):
        for i, d in enumerate(groups[g], start=1):
            setattr(d, "_size_no", f"{g}-{i}")  # 组号-序号
            all_items.append(d)

    if not all_items:
        return jsonify({"success": False, "message": "该项目下无尺寸数据"}), 400

    # 复制并填充
    # 约定：A列 -> SIZE_NO，B列 -> DRAW_DIM（图片优先，无图回退文本渲染PNG；仍失败则写文本）
    def col_letter_to_index(letter: str) -> int:
        letter = letter.upper()
        col = 0
        for ch in letter:
            col = col * 26 + (ord(ch) - ord("A") + 1)
        return col

    colA = col_letter_to_index("A")
    colB = col_letter_to_index("B")

    def fill_one(row_idx: int, item: Any):
        # 复制样式
        _copy_row_style(ws, sample_row, row_idx, c1, c2)
        # 写入 A: SIZE_NO
        ws.cell(row=row_idx, column=colA, value=getattr(item, "_size_no", ""))
        # 写入 B: DRAW_DIM（图片优先）
        placed = False
        # 1) image_url
        img_path = _resolve_asset_to_path(getattr(item, "image_url", "") or "")
        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                anchor = f"{get_column_letter(colB)}{row_idx}"
                ws.add_image(img, anchor)
                placed = True
            except Exception:
                placed = False
        # 2) 文本渲染为 PNG
        if not placed:
            text = _format_dim_text(item)
            export_tmp = os.path.join(_project_root(), "static", "exports", "tmp")
            _ensure_dir(export_tmp)
            tmp_img = os.path.join(export_tmp, f"dim_{uuid.uuid4().hex[:8]}.png")
            png = _render_text_image(text, tmp_img) if text else None
            if png and os.path.exists(png):
                try:
                    img = XLImage(png)
                    anchor = f"{get_column_letter(colB)}{row_idx}"
                    ws.add_image(img, anchor)
                    placed = True
                except Exception:
                    placed = False
        # 3) 仍未放置则写文本
        if not placed:
            ws.cell(row=row_idx, column=colB, value=_format_dim_text(item))

    # 将第一条填在样例行
    fill_one(sample_row, all_items[0])
    # 其余插入到样例行之后
    for idx in range(1, len(all_items)):
        insert_at = sample_row + idx
        ws.insert_rows(insert_at, 1)
        fill_one(insert_at, all_items[idx])

    # 输出目录与文件
    root = _project_root()
    job = uuid.uuid4().hex[:12]
    out_dir = os.path.join(root, "static", "exports", job)
    _ensure_dir(out_dir)
    out_xlsx = os.path.join(out_dir, "filled.xlsx")
    wb.save(out_xlsx)

    # 转换为 HTML（可选）
    html_url = None
    try:
        code, _, _ = _run_soffice_convert(out_xlsx, out_dir, "html")
        if code == 0:
            html_files = glob.glob(os.path.join(out_dir, "*.html"))
            if html_files:
                html_name = os.path.basename(html_files[0])
                html_url = f"/static/exports/{job}/{html_name}"
    except Exception:
        pass

    xlsx_url = f"/static/exports/{job}/filled.xlsx"
    if preview and html_url:
        return redirect(html_url, code=302)

    return jsonify({
        "success": True,
        "job": job,
        "files": {
            "xlsx": xlsx_url,
        "html": html_url,
        }
    })