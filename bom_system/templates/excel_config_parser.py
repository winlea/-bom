#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel配置解析器

解析Excel模板文件的配置信息
"""

import logging
import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelConfigParser:
    """Excel配置解析器"""
    
    def __init__(self):
        self.logger = logger
    
    def parse_template(self, template_path: str) -> Dict[str, Any]:
        """
        解析Excel模板配置
        
        Args:
            template_path: 模板文件路径
            
        Returns:
            解析后的配置信息
        """
        try:
            self.logger.info(f"开始解析Excel模板: {template_path}")
            
            # 加载工作簿
            workbook = load_workbook(template_path, read_only=True)
            
            config = {
                "templatePath": template_path,
                "fileName": os.path.basename(template_path),
                "sheets": [],
                "mainSheet": None,
                "dataStructure": {},
                "projectInfoFields": [],
                "partsTableInfo": {}
            }
            
            # 解析所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = self._analyze_sheet(sheet, sheet_name)
                config["sheets"].append(sheet_info)
                
                # 识别主工作表
                if self._is_main_sheet(sheet, sheet_name):
                    config["mainSheet"] = sheet_name
                    config["dataStructure"] = sheet_info
                    config["projectInfoFields"] = self._extract_project_fields(sheet)
                    config["partsTableInfo"] = self._extract_parts_table_info(sheet)
            
            # 如果没有找到主工作表，使用第一个
            if not config["mainSheet"] and config["sheets"]:
                config["mainSheet"] = config["sheets"][0]["name"]
                config["dataStructure"] = config["sheets"][0]
            
            workbook.close()
            
            self.logger.info(f"Excel模板解析完成，主工作表: {config['mainSheet']}")
            
            return config
            
        except Exception as e:
            self.logger.error(f"解析Excel模板失败: {str(e)}")
            raise
    
    def _analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """分析单个工作表"""
        try:
            sheet_info = {
                "name": sheet_name,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "hasData": False,
                "headerRow": None,
                "dataStartRow": None,
                "columns": [],
                "projectInfoCells": [],
                "type": "unknown"
            }
            
            # 检查是否有数据
            if sheet.max_row > 1:
                sheet_info["hasData"] = True
            
            # 查找表头行
            header_row = self._find_header_row(sheet)
            if header_row:
                sheet_info["headerRow"] = header_row
                sheet_info["dataStartRow"] = header_row + 1
                sheet_info["columns"] = self._extract_column_info(sheet, header_row)
                sheet_info["type"] = "data_table"
            
            # 查找项目信息单元格
            project_cells = self._find_project_info_cells(sheet)
            sheet_info["projectInfoCells"] = project_cells
            
            return sheet_info
            
        except Exception as e:
            self.logger.warning(f"分析工作表 {sheet_name} 失败: {str(e)}")
            return {"name": sheet_name, "error": str(e)}
    
    def _is_main_sheet(self, sheet, sheet_name: str) -> bool:
        """判断是否为主工作表"""
        # 根据工作表名称判断
        main_keywords = ["ODS", "主表", "MAIN", "数据", "零件"]
        if any(keyword in sheet_name.upper() for keyword in main_keywords):
            return True
        
        # 根据内容判断（包含零件相关表头）
        parts_keywords = ["零件编号", "零件名称", "图号", "序号"]
        for row in range(1, min(20, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in cell_value for keyword in parts_keywords):
                        return True
        
        return False
    
    def _find_header_row(self, sheet) -> Optional[int]:
        """查找表头行"""
        header_keywords = ["序号", "零件编号", "零件名称", "图号", "材料"]
        
        for row in range(1, min(20, sheet.max_row + 1)):
            keyword_count = 0
            for col in range(1, min(15, sheet.max_column + 1)):
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if any(keyword in cell_value for keyword in header_keywords):
                        keyword_count += 1
            
            # 如果这一行包含3个或更多关键字，认为是表头行
            if keyword_count >= 3:
                return row
        
        return None
    
    def _extract_column_info(self, sheet, header_row: int) -> List[Dict[str, Any]]:
        """提取列信息"""
        columns = []
        
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(header_row, col)
            if cell.value:
                column_info = {
                    "index": col,
                    "letter": cell.column_letter,
                    "header": str(cell.value).strip(),
                    "fieldName": self._map_header_to_field(str(cell.value).strip()),
                    "dataType": "string"
                }
                columns.append(column_info)
        
        return columns
    
    def _map_header_to_field(self, header: str) -> str:
        """将表头映射到字段名"""
        mappings = {
            "序号": "sequence",
            "零件编号": "part_number",
            "零件名称": "part_name",
            "2D图号": "drawing_2d",
            "3D图号": "drawing_3d",
            "图号": "drawing_2d",
            "原材料": "original_material",
            "终审材料": "final_material_cn",
            "零件分类": "part_category",
            "净重": "net_weight_kg",
            "重量": "net_weight_kg",
            "装配等级": "assembly_level",
            "BOM序号": "bom_sort"
        }
        
        for key, value in mappings.items():
            if key in header:
                return value
        
        # 如果没有匹配，返回处理后的header
        return header.lower().replace(" ", "_").replace("(", "").replace(")", "")
    
    def _find_project_info_cells(self, sheet) -> List[Dict[str, Any]]:
        """查找项目信息单元格"""
        project_keywords = ["项目", "客户", "日期", "版本", "编制", "审核", "批准"]
        project_cells = []
        
        for row in range(1, min(15, sheet.max_row + 1)):
            for col in range(1, min(10, sheet.max_column + 1)):
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    for keyword in project_keywords:
                        if keyword in cell_value:
                            project_cells.append({
                                "row": row,
                                "col": col,
                                "label": cell_value.strip(),
                                "valueCell": {"row": row, "col": col + 1},  # 假设值在右边一列
                                "fieldName": self._map_project_field(cell_value.strip())
                            })
                            break
        
        return project_cells
    
    def _map_project_field(self, label: str) -> str:
        """映射项目字段"""
        mappings = {
            "项目名称": "project_name",
            "项目编号": "project_id",
            "客户": "customer",
            "日期": "date",
            "版本": "version",
            "编制": "author",
            "审核": "reviewer",
            "批准": "approver"
        }
        
        for key, value in mappings.items():
            if key in label:
                return value
        
        return label.lower().replace(" ", "_")
    
    def _extract_project_fields(self, sheet) -> List[Dict[str, Any]]:
        """提取项目信息字段"""
        return self._find_project_info_cells(sheet)
    
    def _extract_parts_table_info(self, sheet) -> Dict[str, Any]:
        """提取零件表信息"""
        header_row = self._find_header_row(sheet)
        
        if not header_row:
            return {}
        
        return {
            "headerRow": header_row,
            "dataStartRow": header_row + 1,
            "columns": self._extract_column_info(sheet, header_row),
            "estimatedDataRows": max(0, sheet.max_row - header_row)
        }
    
    def validate_template(self, template_path: str) -> Dict[str, Any]:
        """
        验证模板文件
        
        Args:
            template_path: 模板文件路径
            
        Returns:
            验证结果
        """
        try:
            validation_result = {
                "isValid": True,
                "errors": [],
                "warnings": [],
                "recommendations": []
            }
            
            # 检查文件是否存在
            if not os.path.exists(template_path):
                validation_result["isValid"] = False
                validation_result["errors"].append("模板文件不存在")
                return validation_result
            
            # 解析配置
            config = self.parse_template(template_path)
            
            # 验证必需字段
            if not config.get("mainSheet"):
                validation_result["warnings"].append("未找到主工作表")
            
            if not config.get("partsTableInfo", {}).get("columns"):
                validation_result["warnings"].append("未找到零件数据列")
            
            # 检查关键列是否存在
            required_fields = ["part_number", "part_name"]
            found_fields = [col["fieldName"] for col in config.get("partsTableInfo", {}).get("columns", [])]
            
            for field in required_fields:
                if field not in found_fields:
                    validation_result["recommendations"].append(f"建议添加 {field} 列")
            
            self.logger.info(f"模板验证完成，有效性: {validation_result['isValid']}")
            
            return validation_result
            
        except Exception as e:
            self.logger.error(f"模板验证失败: {str(e)}")
            return {
                "isValid": False,
                "errors": [f"验证失败: {str(e)}"],
                "warnings": [],
                "recommendations": []
            }

def main():
    """测试配置解析器"""
    template_path = r'templates\WZ1D_Smart_Config_Template.xlsx'
    
    if not os.path.exists(template_path):
        print('❌ 配置模板文件不存在')
        return
    
    print('🔧 Excel配置解析器测试')
    print('=' * 40)
    
    try:
        # 创建解析器
        parser = ExcelConfigParser()
        
        # 加载所有配置
        config = parser.load_all_configurations()
        
        # 打印配置总结
        print('\n📊 配置加载总结:')
        print(f"字段映射: {len(config['data_layer']['field_mappings'])}个")
        print(f"验证规则: {len(config['data_layer']['validation_rules'])}个") 
        print(f"样式映射: {len(config['data_layer']['style_mappings'])}个")
        print(f"模板区域: {len(config['template_layer']['template_regions'])}个")
        print(f"生成规则: {len(config['control_layer']['generation_rules'])}个")
        print(f"异常处理: {len(config['control_layer']['exception_handling'])}个")
        
        # 验证配置
        validation_result = parser.validate_configuration()
        print(f'\n✅ 配置验证: {"通过" if validation_result["is_valid"] else "失败"}')
        
        if validation_result['errors']:
            print('❌ 错误:')
            for error in validation_result['errors']:
                print(f'  - {error}')
        
        if validation_result['warnings']:
            print('⚠️ 警告:')
            for warning in validation_result['warnings']:
                print(f'  - {warning}')
        
        # 导出配置
        output_path = 'exported_config.json'
        if parser.export_config_to_json(output_path):
            print(f'\n📄 配置已导出到: {output_path}')
        
        print('\n🎉 配置解析器测试完成！')
        
    except Exception as e:
        print(f'❌ 测试失败: {e}')

if __name__ == '__main__':
    import os
    main()