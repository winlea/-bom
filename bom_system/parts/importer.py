"""BOM导入解析器 - 支持复杂多Sheet BOM文件"""
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from .models import db, Part, ProjectBomRelation, PartOperation
from bom_system.models import Project
from bom_system.database.session import get_db_session


logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """导入结果"""
    project_created: int = 0
    parts_created: int = 0
    relations_created: int = 0
    operations_created: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# BOM表头映射配置
# Excel列索引 -> 字段名 (0-based)
BOM_COLUMN_MAPPING = {
    0: "序号",
    1: "装配等级",
    8: "U8编码",
    9: "产品编号",
    10: "2D图号",
    11: "3D图号",
    13: "零件名称",
    14: "原图材料",
    15: "终审拟代材料",
    16: "零件分类",
    17: "下料尺寸/规格",
    18: "材料属性",
    19: "业务分类",
}


def parse_bom_file(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """解析BOM文件，返回所有Sheet的数据"""
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    result = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 跳过空白或无关Sheet
        if sheet_name.startswith('Sheet') and not wb[sheet_name].max_row:
            continue
            
        # 读取整个Sheet
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        if data:
            # 转换为DataFrame，保留原始结构
            max_cols = max(len(row) for row in data) if data else 0
            df = pd.DataFrame(data)
            # 补齐列数
            while len(df.columns) < max_cols:
                df[len(df.columns)] = None
            result[sheet_name] = df
    
    return result


def extract_project_info(df: pd.DataFrame) -> Dict[str, str]:
    """从DataFrame提取项目信息"""
    info = {}
    
    # 尝试从第2行提取客户、项目名称等信息
    if len(df) > 1:
        row1 = df.iloc[1].astype(str)
        info['row1_text'] = ' '.join(row1.dropna().tolist())
    
    return info


def find_header_row(df: pd.DataFrame) -> int:
    """查找表头行（包含'序号'、'装配等级'等关键字）"""
    for idx in range(min(10, len(df))):
        row_text = ' '.join(df.iloc[idx].astype(str).tolist())
        if '序号' in row_text and '装配等级' in row_text:
            return idx
    return 3  # 默认第4行（0-based）


def parse_bom_sheet(df: pd.DataFrame, sheet_name: str) -> List[Dict[str, Any]]:
    """解析单个BOM Sheet，提取零件数据"""
    header_row = find_header_row(df)
    parts = []
    
    # 获取表头行
    headers = df.iloc[header_row].tolist()
    
    # 找到关键列的索引
    col_map = {}
    for idx, val in enumerate(headers):
        if pd.isna(val):
            continue
        val_str = str(val).strip()
        for col_name, target in BOM_COLUMN_MAPPING.items():
            if target in val_str:
                col_map[col_name] = idx
    
    logger.info(f"Sheet '{sheet_name}': 表头行={header_row}, 列映射={col_map}")
    
    # 跳过表头，解析数据行
    current_level = 1
    current_sequence = ""
    
    for row_idx in range(header_row + 1, len(df)):
        row = df.iloc[row_idx]
        
        # 获取装配等级
        level_col = col_map.get(1)  # 装配等级
        if level_col is not None:
            level_val = row[level_col]
            if pd.notna(level_val):
                try:
                    current_level = int(float(level_val))
                except (ValueError, TypeError):
                    pass
        
        # 获取序号
        seq_col = col_map.get(0)  # 序号
        if seq_col is not None:
            seq_val = row[seq_col]
            if pd.notna(seq_val):
                current_sequence = str(seq_val).strip()
        
        # 获取零件编号（必需）
        part_col = col_map.get(9)  # 产品编号
        if part_col is None:
            # 尝试扫描U8编码列
            for col_idx in [8, 9, 10]:
                if col_idx in col_map:
                    part_col = col_map[col_idx]
                    break
        
        if part_col is None:
            continue
            
        part_number = row[part_col]
        if pd.isna(part_number):
            continue
        
        part_number = str(part_number).strip()
        if not part_number:
            continue
        
        # 提取零件信息
        part_data = {
            'part_number': part_number,
            'assembly_level': current_level,
            'sequence': current_sequence,
            'source_row': row_idx + 1,  # 1-based
            'source_sheet': sheet_name,
        }
        
        # 提取可选字段
        field_mapping = {
            'u8_code': 8,
            'drawing_2d': 10,
            'drawing_3d': 11,
            'part_name': 13,
            'original_material': 14,
            'final_material_cn': 15,
            'part_category': 16,
            'material_spec': 17,
        }
        
        for field_name, col_idx in field_mapping.items():
            mapped_col = col_map.get(col_idx)
            if mapped_col is not None:
                val = row[mapped_col]
                if pd.notna(val):
                    part_data[field_name] = str(val).strip()
        
        parts.append(part_data)
        logger.debug(f"解析零件: {part_data}")
    
    return parts


def import_bom_file(file_bytes: bytes, project_name: str = None, 
                    filename: str = None) -> ImportResult:
    """导入BOM文件到数据库
    
    Args:
        file_bytes: BOM文件字节数据
        project_name: 项目名称（可选，不传则从文件名推断）
        filename: 原始文件名
    
    Returns:
        ImportResult: 导入结果
    """
    session = get_db_session()
    result = ImportResult()
    
    try:
        # 解析文件
        sheets = parse_bom_file(file_bytes)
        logger.info(f"解析到 {len(sheets)} 个Sheet: {list(sheets.keys())}")
        
        # 确定项目名称
        if not project_name:
            # 从文件名推断
            if filename:
                project_name = filename.replace('.xlsx', '').replace('.xls', '')
            else:
                project_name = "未命名项目"
        
        # 创建或获取项目
        project = session.query(Project).filter_by(name=project_name).first()
        if not project:
            project = Project(name=project_name)
            session.add(project)
            session.flush()
            result.project_created = 1
            logger.info(f"创建项目: {project_name}")
        
        # 处理每个Sheet
        all_parts = []
        for sheet_name, df in sheets.items():
            logger.info(f"处理Sheet: {sheet_name}, 行数: {len(df)}")
            parts = parse_bom_sheet(df, sheet_name)
            for p in parts:
                p['source_sheet'] = sheet_name
            all_parts.extend(parts)
        
        logger.info(f"共解析 {len(all_parts)} 个零件")
        
        # 构建BOM树结构
        bom_sort = 0
        for part_data in all_parts:
            bom_sort += 1
            
            # 获取或创建零件
            part_number = part_data['part_number']
            part = session.query(Part).filter_by(part_number=part_number).first()
            
            is_new_part = False
            if not part:
                part = Part(
                    part_number=part_number,
                    part_name=part_data.get('part_name'),
                    drawing_2d=part_data.get('drawing_2d'),
                    drawing_3d=part_data.get('drawing_3d'),
                    original_material=part_data.get('original_material'),
                    final_material_cn=part_data.get('final_material_cn'),
                    part_category=part_data.get('part_category'),
                    material_spec=part_data.get('material_spec'),
                    u8_code=part_data.get('u8_code'),
                )
                session.add(part)
                session.flush()
                is_new_part = True
                result.parts_created += 1
            
            # 创建项目BOM关系
            relation = ProjectBomRelation(
                project_id=project.id,
                part_id=part.id,
                assembly_level=part_data.get('assembly_level'),
                sequence=part_data.get('sequence'),
                bom_sort=bom_sort,
                source_file=filename,
                source_sheet=part_data.get('source_sheet'),
                source_row=part_data.get('source_row'),
            )
            session.add(relation)
            result.relations_created += 1
        
        session.commit()
        logger.info(f"导入完成: 项目={result.project_created}, 零件={result.parts_created}, 关系={result.relations_created}")
        
    except Exception as e:
        session.rollback()
        result.errors.append(f"导入失败: {str(e)}")
        logger.error(f"导入错误: {e}", exc_info=True)
    
    return result


def get_project_bom_tree(project_id: int) -> List[Dict[str, Any]]:
    """获取项目的BOM树形结构"""
    session = get_db_session()
    
    # 获取项目的所有BOM关系
    relations = session.query(ProjectBomRelation).filter_by(
        project_id=project_id
    ).order_by(ProjectBomRelation.bom_sort).all()
    
    # 构建树形结构
    tree = []
    relation_map = {}  # id -> relation data
    children_map = {}  # parent_id -> [relation_ids]
    
    for rel in relations:
        rel_data = {
            'id': rel.id,
            'part_id': rel.part_id,
            'part_number': rel.part.part_number,
            'part_name': rel.part.part_name,
            'assembly_level': rel.assembly_level,
            'sequence': rel.sequence,
            'quantity': rel.quantity,
            'children': [],
        }
        relation_map[rel.id] = rel_data
        
        if rel.parent_relation_id:
            if rel.parent_relation_id not in children_map:
                children_map[rel.parent_relation_id] = []
            children_map[rel.parent_relation_id].append(rel.id)
        else:
            tree.append(rel_data)
    
    # 填充children
    def fill_children(node):
        children_ids = children_map.get(node['id'], [])
        for cid in children_ids:
            if cid in relation_map:
                child = relation_map[cid]
                node['children'].append(child)
                fill_children(child)
    
    for root in tree:
        fill_children(root)
    
    return tree


def get_part_library(search: str = None, page: int = 1, page_size: int = 50) -> Dict[str, Any]:
    """获取零件库，支持搜索和分页"""
    session = get_db_session()
    
    query = session.query(Part)
    
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            db.or_(
                Part.part_number.like(search_pattern),
                Part.part_name.like(search_pattern),
                Part.drawing_2d.like(search_pattern),
            )
        )
    
    total = query.count()
    parts = query.order_by(Part.part_number).offset((page - 1) * page_size).limit(page_size).all()
    
    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'parts': [
            {
                'id': p.id,
                'part_number': p.part_number,
                'part_name': p.part_name,
                'drawing_2d': p.drawing_2d,
                'drawing_3d': p.drawing_3d,
                'original_material': p.original_material,
                'final_material_cn': p.final_material_cn,
                'part_category': p.part_category,
                'u8_code': p.u8_code,
            }
            for p in parts
        ]
    }
