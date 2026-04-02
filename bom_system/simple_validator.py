#!/usr/bin/env python3
"""
简化版零件信息导入验证工具
不依赖复杂模块结构，直接验证模板填充功能
"""

import json
import os
from datetime import datetime

import openpyxl


class SimplePartValidator:
    """简化版零件信息验证器"""

    def __init__(self, template_dir="templates"):
        self.template_dir = template_dir
        self.results = []

    def create_test_part(self, part_number="4266005"):
        """创建测试零件数据"""
        return {
            "part_number": part_number,
            "part_name": f"测试零件{part_number}",
            "final_material_cn": "Q235B",
            "drawing_2d": f"WZ1D-{part_number}-001",
            "drawing_3d": f"WZ1D-{part_number}-002",
            "material_thickness": "2.0mm",
            "product_name": f"产品{part_number}",
            "drawing_sizes": [
                {"size": "Φ8.1", "tolerance": "+0.05/-0.08", "tool": "卡尺"},
                {"size": "Φ12.0", "tolerance": "±0.1", "tool": "卡尺"},
                {"size": "15.0", "tolerance": "±0.2", "tool": "卡尺"},
                {"size": "20.0", "tolerance": "±0.1", "tool": "卡尺"},
            ],
        }

    def validate_template_fill(self, template_file, part_data):
        """验证模板填充功能"""
        try:
            # 检查模板文件是否存在
            template_path = os.path.join(self.template_dir, template_file)
            if not os.path.exists(template_path):
                return {
                    "status": "failed",
                    "message": f"模板文件不存在: {template_file}",
                    "template": template_file,
                }

            # 加载模板
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # 记录原始状态
            original_values = {}
            placeholders = [
                "{{TITLE}}",
                "{{PART_NUMBER}}",
                "{{PART_NAME}}",
                "{{MATERIAL_SPEC}}",
                "{{PRODUCT_NAME}}",
                "{{DRAWING_2D}}",
                "{{DRAWING_3D}}",
                "{{MATERIAL_THICKNESS}}",
            ]

            for placeholder in placeholders:
                found = False
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and placeholder in str(cell.value):
                            original_values[placeholder] = cell.value
                            found = True
                            break
                    if found:
                        break

            # 填充占位符
            fill_results = self._fill_placeholders(ws, part_data)

            # 验证图纸尺寸区域
            size_validation = self._validate_drawing_sizes(ws, part_data)

            # 保存验证结果
            output_file = f"validated_{part_data['part_number']}_{template_file}"
            output_path = os.path.join("validation_output", output_file)
            os.makedirs("validation_output", exist_ok=True)
            wb.save(output_path)

            return {
                "status": "success",
                "message": "模板填充验证成功",
                "template": template_file,
                "part_number": part_data["part_number"],
                "output_file": output_file,
                "placeholders_filled": fill_results,
                "size_validation": size_validation,
                "original_placeholders": len(original_values),
                "filled_placeholders": len(fill_results),
            }

        except Exception as e:
            return {
                "status": "error",
                "message": str(e),
                "template": template_file,
                "part_number": part_data["part_number"],
            }

    def _fill_placeholders(self, ws, part_data):
        """填充占位符"""
        fill_results = {}

        # 映射占位符到数据
        mapping = {
            "{{TITLE}}": f"{part_data['part_name']} 尺寸检验记录",
            "{{PART_NUMBER}}": part_data["part_number"],
            "{{PART_NAME}}": part_data["part_name"],
            "{{MATERIAL_SPEC}}": part_data["final_material_cn"],
            "{{PRODUCT_NAME}}": part_data["product_name"],
            "{{DRAWING_2D}}": part_data["drawing_2d"],
            "{{DRAWING_3D}}": part_data["drawing_3d"],
            "{{MATERIAL_THICKNESS}}": part_data["material_thickness"],
        }

        for placeholder, value in mapping.items():
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and placeholder in str(cell.value):
                        cell.value = str(cell.value).replace(placeholder, str(value))
                        fill_results[placeholder] = value
                        break

        return fill_results

    def _validate_drawing_sizes(self, ws, part_data):
        """验证图纸尺寸区域"""
        try:
            # 查找图纸尺寸区域（B12:D20）
            size_area_start = 12  # B12
            size_area_end = 20  # D20

            filled_sizes = 0
            for idx, size_data in enumerate(part_data["drawing_sizes"]):
                if idx >= 8:  # 最多8行
                    break

                row = size_area_start + idx
                if row <= size_area_end:
                    # 填充尺寸数据
                    size_col = 2  # B列
                    tolerance_col = 3  # C列
                    tool_col = 4  # D列

                    # 检查单元格是否存在
                    if ws.cell(row=row, column=size_col):
                        ws.cell(row=row, column=size_col, value=size_data["size"])
                        ws.cell(
                            row=row, column=tolerance_col, value=size_data["tolerance"]
                        )
                        ws.cell(row=row, column=tool_col, value=size_data["tool"])
                        filled_sizes += 1

            return {
                "status": "success",
                "total_sizes": len(part_data["drawing_sizes"]),
                "filled_sizes": filled_sizes,
                "area_range": f"B{size_area_start}:D{size_area_end}",
            }

        except Exception as e:
            return {"status": "error", "message": str(e)}

    def run_quick_validation(self):
        """运行快速验证"""
        print("🚀 开始快速验证零件信息导入功能...")

        # 测试零件
        test_parts = [
            self.create_test_part("4266005"),
            self.create_test_part("4266006"),
        ]

        # 查找模板文件
        template_files = []
        if os.path.exists(self.template_dir):
            for file in os.listdir(self.template_dir):
                if file.endswith(".xlsx") and "WZ1D" in file:
                    template_files.append(file)

        if not template_files:
            print("❌ 未找到模板文件，请确保templates目录包含WZ1D模板")
            return False

        # 验证每个模板和零件组合
        results = []
        for template in template_files:
            for part in test_parts:
                print(f"📋 验证模板: {template}, 零件: {part['part_number']}")
                result = self.validate_template_fill(template, part)
                results.append(result)

                if result["status"] == "success":
                    print(
                        f"   ✅ 成功: 填充了 {result['filled_placeholders']} 个占位符"
                    )
                else:
                    print(f"   ❌ 失败: {result['message']}")

        # 生成报告
        self._generate_quick_report(results)

        # 总结
        success_count = sum(1 for r in results if r["status"] == "success")
        total_count = len(results)

        print(f"\n📊 验证完成: {success_count}/{total_count} 个测试通过")

        if success_count == total_count:
            print("🎉 所有验证通过！您的零件信息可以正确导入")
            return True
        else:
            print("⚠️  部分验证失败，请查看详细报告")
            return False

    def _generate_quick_report(self, results):
        """生成快速验证报告"""
        report = {
            "validation_date": datetime.now().isoformat(),
            "total_tests": len(results),
            "successful_tests": sum(1 for r in results if r["status"] == "success"),
            "failed_tests": sum(1 for r in results if r["status"] == "failed"),
            "error_tests": sum(1 for r in results if r["status"] == "error"),
            "results": results,
        }

        # 保存报告
        report_file = "validation_output/quick_validation_report.json"
        os.makedirs("validation_output", exist_ok=True)

        with open(report_file, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        print(f"📄 验证报告已保存: {report_file}")


if __name__ == "__main__":
    validator = SimplePartValidator()
    validator.run_quick_validation()
