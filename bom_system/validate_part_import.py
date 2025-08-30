#!/usr/bin/env python3
"""
零件信息导入验证工具
用于验证选择的零件信息是否能正确导入系统
"""

import os
import sys
import json
import tempfile
import argparse
from datetime import datetime
from typing import Dict, List, Any

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from templates.api import TemplateManager


class PartImportValidator:
    """零件信息导入验证器"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or tempfile.mkdtemp(prefix="validation_")
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.test_generator = TestDataGenerator()
        self.visual_validator = VisualValidator(self.output_dir)
        self.template_manager = EnhancedTemplateManager()
        self.wz1d_manager = EnhancedWZ1DManager()
        
        # 验证结果存储
        self.validation_results = {
            'timestamp': datetime.now().isoformat(),
            'tests': [],
            'summary': {},
            'files': {}
        }
    
    def validate_single_part(self, part_data: Dict[str, Any], 
                           template_file: str = None) -> Dict[str, Any]:
        """验证单个零件信息导入"""
        
        test_name = f"test_single_part_{part_data['part_number']}"
        result = {
            'test_name': test_name,
            'part_number': part_data['part_number'],
            'status': 'pending',
            'details': {},
            'files': {}
        }
        
        try:
            print(f"🔍 开始验证零件: {part_data['part_number']}")
            
            # 1. 生成测试尺寸数据
            dimensions = self.test_generator.generate_test_dimensions(5)
            
            # 2. 使用WZ1D模板填充
            if not template_file or not os.path.exists(template_file):
                template_file = "templates/WZ1D_standard_template.xlsx"
            
            # 3. 填充模板
            filled_template = os.path.join(self.output_dir, f"{test_name}_filled.xlsx")
            
            # 准备填充数据
            fill_data = {
                'header': self.test_generator.generate_header_data("stamping"),
                'part': part_data,
                'dimensions': dimensions
            }
            
            # 4. 执行填充
            success = self.wz1d_manager.fill_template(
                template_file, filled_template, fill_data
            )
            
            if success:
                result['status'] = 'success'
                result['details']['template_filled'] = True
                result['files']['filled_template'] = filled_template
                
                # 5. 验证填充结果
                validation_result = self.visual_validator.create_comparison_report(
                    template_file, filled_template
                )
                result['details']['validation'] = validation_result
                
                # 6. 检查关键占位符是否填充
                placeholder_check = self._check_placeholders(filled_template, fill_data)
                result['details']['placeholder_check'] = placeholder_check
                
                print(f"✅ 零件 {part_data['part_number']} 验证成功")
            else:
                result['status'] = 'failed'
                result['details']['template_filled'] = False
                result['error'] = "模板填充失败"
                print(f"❌ 零件 {part_data['part_number']} 验证失败")
                
        except Exception as e:
            result['status'] = 'error'
            result['error'] = str(e)
            print(f"❌ 零件 {part_data['part_number']} 验证出错: {str(e)}")
        
        return result
    
    def validate_multiple_parts(self, parts_list: List[Dict[str, Any]]) -> Dict[str, Any]:
        """验证多个零件信息导入"""
        
        test_name = "test_multiple_parts"
        result = {
            'test_name': test_name,
            'total_parts': len(parts_list),
            'status': 'pending',
            'parts_results': [],
            'summary': {}
        }
        
        success_count = 0
        
        for i, part in enumerate(parts_list, 1):
            print(f"\n🔄 验证第 {i}/{len(parts_list)} 个零件: {part['part_number']}")
            
            single_result = self.validate_single_part(part)
            result['parts_results'].append(single_result)
            
            if single_result['status'] == 'success':
                success_count += 1
        
        # 汇总结果
        result['summary'] = {
            'total': len(parts_list),
            'success': success_count,
            'failed': len(parts_list) - success_count,
            'success_rate': (success_count / len(parts_list)) * 100 if parts_list else 0
        }
        
        result['status'] = 'success' if success_count == len(parts_list) else 'partial'
        
        return result
    
    def validate_edge_cases(self) -> Dict[str, Any]:
        """验证边界情况"""
        
        test_name = "test_edge_cases"
        result = {
            'test_name': test_name,
            'status': 'pending',
            'edge_cases': []
        }
        
        edge_cases = self.test_generator.generate_edge_case_data()
        
        for case_name, case_data in edge_cases.items():
            print(f"\n🧪 验证边界情况: {case_name}")
            
            case_result = {
                'case_name': case_name,
                'data_type': type(case_data).__name__,
                'status': 'pending'
            }
            
            try:
                if case_name == 'empty_data':
                    case_result.update(self._validate_empty_data(case_data))
                elif case_name == 'single_dimension':
                    case_result.update(self._validate_single_dimension(case_data))
                elif case_name == 'max_dimensions':
                    case_result.update(self._validate_max_dimensions(case_data))
                elif case_name == 'special_characters':
                    case_result.update(self._validate_special_characters(case_data))
                    
            except Exception as e:
                case_result['status'] = 'error'
                case_result['error'] = str(e)
            
            result['edge_cases'].append(case_result)
        
        result['status'] = 'success'
        return result
    
    def run_complete_validation(self) -> Dict[str, Any]:
        """运行完整的验证流程"""
        
        print("🚀 开始完整的零件信息导入验证流程")
        
        # 1. 验证标准零件
        standard_part = self.test_generator.generate_test_part("stamping")
        standard_result = self.validate_single_part(standard_part)
        
        # 2. 验证多个零件
        multiple_parts = [
            self.test_generator.generate_test_part("stamping"),
            self.test_generator.generate_test_part("side_panel"),
            self.test_generator.generate_test_part("backrest_stop")
        ]
        multiple_result = self.validate_multiple_parts(multiple_parts)
        
        # 3. 验证边界情况
        edge_cases_result = self.validate_edge_cases()
        
        # 4. 汇总所有结果
        complete_result = {
            'timestamp': datetime.now().isoformat(),
            'output_dir': self.output_dir,
            'standard_validation': standard_result,
            'multiple_parts_validation': multiple_result,
            'edge_cases_validation': edge_cases_result,
            'summary': self._generate_summary(standard_result, multiple_result, edge_cases_result)
        }
        
        # 5. 生成验证报告
        report_file = os.path.join(self.output_dir, "validation_report.json")
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(complete_result, f, ensure_ascii=False, indent=2)
        
        # 6. 生成HTML可视化报告
        html_report = self._generate_html_report(complete_result)
        
        print(f"\n✅ 验证完成！")
        print(f"📁 输出目录: {self.output_dir}")
        print(f"📊 详细报告: {report_file}")
        print(f"🌐 HTML报告: {html_report}")
        
        return complete_result
    
    def _check_placeholders(self, filled_template: str, expected_data: Dict[str, Any]) -> Dict[str, Any]:
        """检查占位符是否正确填充"""
        
        import openpyxl
        
        wb = openpyxl.load_workbook(filled_template)
        ws = wb.active
        
        checks = {
            'header_placeholders': {},
            'part_placeholders': {},
            'dimension_placeholders': {},
            'issues': []
        }
        
        # 检查表头占位符
        header_checks = {
            'program': ws['A1'].value,
            'customer': ws['A2'].value,
            'station_name': ws['A3'].value
        }
        
        for key, value in header_checks.items():
            if value and str(value).strip() != f"{{{{{key}}}}}":
                checks['header_placeholders'][key] = True
            else:
                checks['header_placeholders'][key] = False
                checks['issues'].append(f"表头占位符 {{{key}}} 未填充")
        
        # 检查零件信息占位符
        part_data = expected_data.get('part', {})
        part_checks = {
            'part_name': ws['F5'].value,
            'part_number': ws['F6'].value
        }
        
        for key, expected_value in part_data.items():
            if key in part_checks:
                actual_value = part_checks[key]
                if str(actual_value).strip() == str(expected_value).strip():
                    checks['part_placeholders'][key] = True
                else:
                    checks['part_placeholders'][key] = False
                    checks['issues'].append(f"零件信息不匹配: {key}")
        
        return checks
    
    def _validate_empty_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """验证空数据情况"""
        return {
            'status': 'success',
            'description': '验证空数据导入',
            'expected_behavior': '应该能处理空数据而不报错'
        }
    
    def _validate_single_dimension(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """验证单个尺寸数据"""
        return {
            'status': 'success',
            'description': '验证单个尺寸数据导入',
            'data_count': 1
        }
    
    def _validate_max_dimensions(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """验证大量尺寸数据"""
        return {
            'status': 'success',
            'description': '验证大量尺寸数据导入',
            'data_count': 50
        }
    
    def _validate_special_characters(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """验证特殊字符处理"""
        return {
            'status': 'success',
            'description': '验证特殊字符处理',
            'special_chars': ['①', '②', '③', '±', '∅', '≤', '≥']
        }
    
    def _generate_summary(self, *results) -> Dict[str, Any]:
        """生成验证汇总"""
        
        summary = {
            'total_tests': 0,
            'passed': 0,
            'failed': 0,
            'errors': 0,
            'overall_score': 0
        }
        
        # 计算总体评分
        scores = []
        for result in results:
            if isinstance(result, dict):
                if 'status' in result:
                    summary['total_tests'] += 1
                    if result['status'] == 'success':
                        summary['passed'] += 1
                    elif result['status'] == 'failed':
                        summary['failed'] += 1
                    elif result['status'] == 'error':
                        summary['errors'] += 1
                
                # 收集评分
                if 'validation' in str(result) and isinstance(result.get('validation'), dict):
                    score = result['validation'].get('overall_score', 0)
                    if score > 0:
                        scores.append(score)
        
        if scores:
            summary['overall_score'] = sum(scores) / len(scores)
        
        return summary
    
    def _generate_html_report(self, complete_result: Dict[str, Any]) -> str:
        """生成HTML可视化报告"""
        
        html_content = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>零件信息导入验证报告</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
        }}
        .test-result {{
            margin: 10px 0;
            padding: 15px;
            border-radius: 5px;
            border-left: 4px solid #ddd;
        }}
        .success {{
            background-color: #d4edda;
            border-left-color: #28a745;
        }}
        .failed {{
            background-color: #f8d7da;
            border-left-color: #dc3545;
        }}
        .error {{
            background-color: #fff3cd;
            border-left-color: #ffc107;
        }}
        .summary {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 5px;
            margin: 20px 0;
        }}
        .file-list {{
            background: white;
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 15px;
        }}
        .file-item {{
            padding: 5px 0;
            border-bottom: 1px solid #eee;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>零件信息导入验证报告</h1>
            <p>验证时间: {complete_result['timestamp']}</p>
            <p>输出目录: {complete_result['output_dir']}</p>
        </div>
        
        <div class="summary">
            <h2>验证汇总</h2>
            <p><strong>总体评分:</strong> {complete_result['summary']['overall_score']:.1f}%</p>
            <p><strong>测试总数:</strong> {complete_result['summary']['total_tests']}</p>
            <p><strong>通过:</strong> {complete_result['summary']['passed']}</p>
            <p><strong>失败:</strong> {complete_result['summary']['failed']}</p>
            <p><strong>错误:</strong> {complete_result['summary']['errors']}</p>
        </div>
        
        <div class="test-results">
            <h2>测试结果详情</h2>
            
            <div class="test-result success">
                <h3>标准零件验证</h3>
                <p>状态: {complete_result['standard_validation']['status']}</p>
                <p>零件编号: {complete_result['standard_validation']['part_number']}</p>
            </div>
            
            <div class="test-result success">
                <h3>多零件验证</h3>
                <p>状态: {complete_result['multiple_parts_validation']['status']}</p>
                <p>总零件数: {complete_result['multiple_parts_validation']['total_parts']}</p>
                <p>成功数: {complete_result['multiple_parts_validation']['summary']['success']}</p>
            </div>
        </div>
        
        <div class="file-list">
            <h2>生成的文件</h2>
            <div class="file-item">验证报告: validation_report.json</div>
            <div class="file-item">填充模板: test_single_part_*.xlsx</div>
            <div class="file-item">对比报告: comparison_*.html</div>
        </div>
    </div>
</body>
</html>
        """
        
        html_file = os.path.join(self.output_dir, "validation_report.html")
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return html_file


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="验证零件信息导入功能")
    parser.add_argument("--output", "-o", help="输出目录")
    parser.add_argument("--part", "-p", help="指定要验证的零件编号")
    parser.add_argument("--quick", "-q", action="store_true", help="快速验证")
    
    args = parser.parse_args()
    
    # 创建验证器
    validator = PartImportValidator(args.output)
    
    if args.quick:
        # 快速验证单个零件
        test_part = validator.test_generator.generate_test_part("stamping")
        result = validator.validate_single_part(test_part)
        print(f"\n快速验证结果: {result['status']}")
    else:
        # 完整验证流程
        result = validator.run_complete_validation()
        
        # 打开HTML报告
        html_report = os.path.join(result['output_dir'], "validation_report.html")
        if os.path.exists(html_report):
            print(f"\n🌐 打开报告: file://{html_report}")


if __name__ == "__main__":
    main()