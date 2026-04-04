"""ODS Controller - Handles ODS preview and generation"""
import logging
import os
import tempfile
from io import BytesIO

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from sqlalchemy import or_, and_

from ..response import APIResponse
from ...database.session import get_db_session
from ...models import BomTable
from ...dimensions.models import Dimension
from ...dimensions.services import DimensionService
from ...dimensions.formatters import format_dimension

logger = logging.getLogger(__name__)
ods_bp = Blueprint("ods", __name__)


def _get_part_dimensions(session, part):
    """Get dimensions for a part from database"""
    part_number = part.part_number
    project_id = part.project_id
    
    logger.debug(f"查询尺寸数据: part_id={part.id}, part_number={part_number}, project_id={project_id}")
    
    service = DimensionService(session)
    
    dimension_records = service.get_dimensions_by_part_number(project_id, part_number)
    logger.debug(f"通过零件编号{part_number}查询到 {len(dimension_records)} 条记录")
    
    if not dimension_records:
        part_id_str = str(part.id)
        dimension_records = service.get_dimensions_by_part(part_id_str)
        logger.debug(f"通过part_id={part_id_str}查询到 {len(dimension_records)} 条记录")
    
    related_image_dimensions = session.query(Dimension).filter(
        or_(
            and_(
                Dimension.dimension_type.in_(['image', 'image_dimension']),
                Dimension.part_id == part_number
            ),
            and_(
                Dimension.dimension_type.in_(['image', 'image_dimension']),
                Dimension.part_id == ""
            )
        )
    ).all()
    logger.debug(f"找到 {len(related_image_dimensions)} 条与当前零件相关的图片尺寸记录")
    
    all_dimensions = dimension_records.copy()
    existing_ids = {dim.id for dim in dimension_records}
    for img_dim in related_image_dimensions:
        if img_dim.id not in existing_ids:
            all_dimensions.append(img_dim)
    
    if not all_dimensions:
        all_dimensions = service.get_dimensions_by_project(project_id)
        logger.debug(f"通过项目ID{project_id}查询到 {len(all_dimensions)} 条记录")
    
    return all_dimensions


@ods_bp.route("/ods/preview/<int:part_id>", methods=["GET"])
def get_ods_preview(part_id):
    """Get ODS preview data"""
    try:
        session = get_db_session()
        part = session.query(BomTable).get(part_id)
        
        if part:
            dimension_records = _get_part_dimensions(session, part)
        else:
            logger.debug(f"未找到零件: part_id={part_id}")
            dimension_records = []
        
        dimensions = []
        for dim in dimension_records:
            technical_note = ""
            if dim.dimension_type in ['image', 'image_dimension']:
                technical_note = "图片尺寸"
            else:
                technical_note = format_dimension(dim)
            
            dimensions.append({
                "sequence_no": dim.id,
                "technical_note": technical_note,
                "measurement_method": "卡尺/游标卡尺",
                "characteristic_code": dim.characteristic or "",
                "frequency": "首/巡/末检",
                "image_url": dim.image_url or ""
            })
        
        if not dimensions:
            dimensions = [
                {"sequence_no": 1, "technical_note": "尺寸1: 100±0.1", "measurement_method": "卡尺", "characteristic_code": "CR", "frequency": "首/巡/末检"},
                {"sequence_no": 2, "technical_note": "尺寸2: 50±0.05", "measurement_method": "游标卡尺", "characteristic_code": "MA", "frequency": "首/巡/末检"}
            ]
        
        return APIResponse.success(data={"dimensions": dimensions, "part_id": part_id})
    except Exception as e:
        logger.debug(f"获取ODS预览数据失败: {str(e)}")
        return APIResponse.internal_server_error(
            message=f"Get ODS preview failed: {str(e)}",
            errors={"error": "get ods preview failed"}
        )


@ods_bp.route("/ods/generate", methods=["POST"])
def generate_ods():
    """Generate ODS inspection guide document"""
    try:
        logger.debug("开始生成ODS")
        data = request.get_json()
        logger.debug(f"获取到数据: {data}")
        
        if not data:
            return APIResponse.bad_request(message="request body required")

        part_id = data.get("part_id")
        if not part_id:
            return APIResponse.bad_request(message="part_id is required")

        try:
            part_id = int(part_id)
        except ValueError:
            return APIResponse.bad_request(message="part_id must be an integer")

        session = get_db_session()
        part = session.query(BomTable).get(part_id)
        logger.debug(f"获取到零件信息: {part}")
        
        if not part:
            return APIResponse.not_found(message="part not found")

        project = session.query(BomTable.__bases__[0].__subclasses__()[0] if hasattr(BomTable, '__bases__') else None).get(part.project_id)
        from ...models import Project
        project = session.query(Project).get(part.project_id)
        logger.debug(f"获取到项目信息: {project}")
        
        if not project:
            return APIResponse.not_found(message="project not found")

        project_info = {
            "id": project.id,
            "name": project.name,
            "customer_name": project.customer_name,
            "description": project.description,
        }
        
        dimension_records = _get_part_dimensions(session, part)
        
        dimensions = []
        for dim in dimension_records:
            drawing = ""
            if hasattr(dim, 'dimension_type'):
                if dim.dimension_type in ['image', 'image_dimension']:
                    drawing = "图片尺寸"
                else:
                    drawing = format_dimension(dim)
            
            dimensions.append({
                "drawing": drawing,
                "method": "卡尺/游标卡尺",
                "special": dim.characteristic or "" if hasattr(dim, 'characteristic') else "",
                "frequency": "首/巡/末检",
                "imageUrl": dim.image_url or "" if hasattr(dim, 'image_url') else ""
            })
        
        if not dimensions:
            dimensions = [
                {"drawing": "尺寸1: 100±0.1", "method": "卡尺", "special": "CR", "frequency": "首/巡/末检"},
                {"drawing": "尺寸2: 50±0.05", "method": "游标卡尺", "special": "MA", "frequency": "首/巡/末检"}
            ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ODS数据"
        
        sheet["A1"] = f"项目：{project_info.get('name', '')}"
        sheet["A2"] = f"客户：{project_info.get('customer_name', '')}"
        
        headers = ["序号", "特征描述", "检测方法", "特性号码", "频率"]
        for col, header in enumerate(headers, 1):
            sheet.cell(4, col).value = header
        
        for row, dim in enumerate(dimensions, 5):
            sheet.cell(row, 1).value = row - 4
            sheet.cell(row, 2).value = dim.get("drawing", "")
            sheet.cell(row, 3).value = dim.get("method", "")
            sheet.cell(row, 4).value = dim.get("special", "")
            sheet.cell(row, 5).value = dim.get("frequency", "")
        
        output_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        output_file.close()
        workbook.save(output_file.name)

        return send_file(
            output_file.name,
            as_attachment=True,
            download_name=f"ODS_{part.part_name}_{part.part_number}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.debug(f"ODS generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        return APIResponse.internal_server_error(
            message=f"Generate ODS failed: {str(e)}",
            errors={"error": "generate ODS failed"}
        )
