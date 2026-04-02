#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修正版多尺寸验证工具
使用用户提供的真实尺寸数据
"""

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side


class CorrectDimensionValidator:
    """修正后的尺寸验证器，使用真实数据"""

    def __init__(self):
        self.output_dir = "validation_output"
        self.ensure_output_dir()

        # 用户提供的真实尺寸数据
        self.real_dimensions = {
            "4266010": [
                {"name": "DIM_1", "value": 87.24, "unit": "mm", "tolerance": "±0.21"},
                {"name": "DIM_2", "value": 103.43, "unit": "mm", "tolerance": "±0.14"},
                {"name": "DIM_3", "value": 134.38, "unit": "mm", "tolerance": "±0.32"},
                {"name": "DIM_4", "value": 179.61, "unit": "mm", "tolerance": "±0.26"},
                {"name": "DIM_5", "value": 29.58, "unit": "mm", "tolerance": "±0.5"},
                {"name": "DIM_6", "value": 135.93, "unit": "mm", "tolerance": "±0.38"},
                {"name": "DIM_7", "value": 192.21, "unit": "mm", "tolerance": "±0.13"},
                {"name": "DIM_8", "value": 177.42, "unit": "mm", "tolerance": "±0.47"},
                {"name": "DIM_9", "value": 36.45, "unit": "mm", "tolerance": "±0.21"},
                {"name": "DIM_10", "value": 132.05, "unit": "mm", "tolerance": "±0.11"},
            ]
        }

    def ensure_output_dir(self):
        """确保输出目录存在"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def create_corrected_template(self, part_number):
        """创建修正后的模板，使用真实尺寸数据"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "零件信息"

        # 设置列宽
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20

        # 零件基础信息
        ws["A1"] = "零件信息"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")

        ws["A3"] = "零件编号"
        ws["B3"] = part_number
        ws["A4"] = "零件名称"
        ws["B4"] = f"测试零件_{part_number}"
        ws["A5"] = "类型"
        ws["B5"] = "MULTI_DIMENSION_TEST"

        # 尺寸信息标题
        ws["A7"] = "尺寸信息"
        ws["A7"].font = Font(bold=True, size=12)
        ws.merge_cells("A7:F7")

        # 图纸尺寸标题
        ws["A9"] = "图纸尺寸"
        ws["A9"].font = Font(bold=True, size=12)
        ws.merge_cells("A9:F9")

        # 表头 - 从B12开始
        ws["B11"] = "尺寸名称"
        ws["C11"] = "数值"
        ws["D11"] = "单位"
        ws["E11"] = "公差"

        for col in ["B11", "C11", "D11", "E11"]:
            ws[col].font = Font(bold=True)
            ws[col].alignment = Alignment(horizontal="center")

        # 填充真实尺寸数据 - 从B12开始
        dimensions = self.real_dimensions.get(part_number, [])
        for i, dim in enumerate(dimensions, start=12):
            ws[f"B{i}"] = dim["name"]
            ws[f"C{i}"] = dim["value"]
            ws[f"D{i}"] = dim["unit"]
            ws[f"E{i}"] = dim["tolerance"]

            # 设置样式
            ws[f"C{i}"].number_format = "0.00"
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{i}"].alignment = Alignment(horizontal="center")

        # 保存文件
        filename = (
            f"corrected_{part_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        return filepath, len(dimensions)

    def validate_corrected_data(self, filepath, part_number):
        """验证修正后的数据准确性"""
        wb = openpyxl.load_workbook(filepath)
        ws = wb["零件信息"]

        errors = []

        # 验证零件信息
        actual_part_number = ws["B3"].value
        if str(actual_part_number) != str(part_number):
            errors.append(
                f"零件编号不匹配: 期望 {part_number}, 实际 {actual_part_number}"
            )

        # 验证尺寸数据
        expected_dims = self.real_dimensions.get(part_number, [])
        actual_dims = []

        row = 12
        while ws[f"B{row}"].value is not None:
            dim_name = ws[f"B{row}"].value
            dim_value = ws[f"C{row}"].value
            dim_unit = ws[f"D{row}"].value
            dim_tolerance = ws[f"E{row}"].value

            actual_dims.append(
                {
                    "name": dim_name,
                    "value": float(dim_value) if dim_value is not None else 0,
                    "unit": str(dim_unit) if dim_unit else "",
                    "tolerance": str(dim_tolerance) if dim_tolerance else "",
                }
            )
            row += 1

        # 验证尺寸数量和数值
        if len(actual_dims) != len(expected_dims):
            errors.append(
                f"尺寸数量不匹配: 期望 {len(expected_dims)}, 实际 {len(actual_dims)}"
            )

        # 验证具体数值
        for i, (expected, actual) in enumerate(zip(expected_dims, actual_dims)):
            if abs(expected["value"] - actual["value"]) > 0.01:
                errors.append(
                    f"尺寸 {expected['name']} 数值不匹配: 期望 {expected['value']}, 实际 {actual['value']}"
                )

            if expected["unit"] != actual["unit"]:
                errors.append(
                    f"尺寸 {expected['name']} 单位不匹配: 期望 {expected['unit']}, 实际 {actual['unit']}"
                )

            if expected["tolerance"] != actual["tolerance"]:
                errors.append(
                    f"尺寸 {expected['name']} 公差不匹配: 期望 {expected['tolerance']}, 实际 {actual['tolerance']}"
                )

        return {
            "is_valid": len(errors) == 0,
            "errors": errors,
            "expected_count": len(expected_dims),
            "actual_count": len(actual_dims),
            "dimensions": actual_dims,
        }

    def run_correction_test(self):
        """运行修正后的测试"""
        print("运行修正后的尺寸验证测试...")

        results = []

        for part_number in self.real_dimensions.keys():
            print(f"验证零件 {part_number}...")

            # 创建修正后的模板
            filepath, dim_count = self.create_corrected_template(part_number)

            # 验证数据
            validation = self.validate_corrected_data(filepath, part_number)

            result = {
                "part_number": part_number,
                "filepath": filepath,
                "expected_dimensions": len(self.real_dimensions[part_number]),
                "actual_dimensions": validation["actual_count"],
                "is_valid": validation["is_valid"],
                "errors": validation["errors"],
                "dimensions_data": self.real_dimensions[part_number],
            }

            results.append(result)

            if validation["is_valid"]:
                print(f"✓ 零件 {part_number} 验证通过")
            else:
                print(f"✗ 零件 {part_number} 验证失败: {validation['errors']}")

        # 生成测试报告
        report = {
            "test_name": "修正后尺寸验证测试",
            "test_date": datetime.now().isoformat(),
            "total_tests": len(results),
            "passed_tests": sum(1 for r in results if r["is_valid"]),
            "failed_tests": sum(1 for r in results if not r["is_valid"]),
            "results": results,
            "real_dimensions_used": self.real_dimensions,
        }

        report_path = os.path.join(
            self.output_dir, "corrected_dimension_test_report.json"
        )
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        print(f"\n修正后测试完成:")
        print(f"总测试数: {report['total_tests']}")
        print(f"通过测试: {report['passed_tests']}")
        print(f"失败测试: {report['failed_tests']}")
        print(f"详细报告: {report_path}")

        return results


if __name__ == "__main__":
    validator = CorrectDimensionValidator()
    validator.run_correction_test()
