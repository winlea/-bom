"""
尺寸管理API
"""

import logging

from flask import Blueprint, jsonify, request

from bom_system.database.session import get_db_session
from bom_system.dimensions.quality_service import QualityControlService
from bom_system.dimensions.services import DimensionService, DimensionValidationError

# 创建蓝图
dimensions_bp = Blueprint("dimensions", __name__, url_prefix="/api/dimensions")

# 配置日志
logger = logging.getLogger(__name__)


@dimensions_bp.route("", methods=["GET"])
def list_dimensions():
    """通用尺寸列表：优先按 partId 过滤，其次按 projectId"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        part_id = request.args.get("partId")
        project_id = request.args.get("projectId")
        search_term = (request.args.get("search") or "").strip()

        dims = []
        if part_id:
            dims = service.get_dimensions_by_part(part_id)
            if search_term:
                dims = [
                    d
                    for d in dims
                    if (d.characteristic or "").lower().find(search_term.lower()) != -1
                ]
        elif project_id:
            if search_term:
                dims = service.search_dimensions(project_id, search_term)
            else:
                dims = service.get_dimensions_by_project(project_id)
        else:
            return jsonify({"success": False, "message": "请提供 partId 或 projectId"}), 400

        result = [d.to_dict() for d in dims]
        return jsonify({"success": True, "data": result, "total": len(result)})
    except Exception as e:
        logger.error(f"查询尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"查询尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/projects/<project_id>", methods=["GET"])
def get_project_dimensions(project_id):
    """获取项目的尺寸数据，支持按产品编号过滤"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        # 获取查询参数
        search_term = request.args.get("search", "").strip()
        part_number = request.args.get("part_number", "").strip()

        # 优先按产品编号过滤，其次按搜索词，最后返回项目所有尺寸
        if part_number:
            # 根据产品编号获取该零件的尺寸
            dimensions = service.get_dimensions_by_part_number(project_id, part_number)
            if search_term:
                # 在零件尺寸中进一步搜索
                dimensions = [
                    d
                    for d in dimensions
                    if (d.characteristic or "").lower().find(search_term.lower()) != -1
                ]
        elif search_term:
            dimensions = service.search_dimensions(project_id, search_term)
        else:
            dimensions = service.get_dimensions_by_project(project_id)

        # 转换为字典格式，确保字段名匹配前端期望
        result = []
        for dim in dimensions:
            dim_dict = dim.to_dict()
            # 确保字段名匹配前端期望的格式
            result.append(
                {
                    "id": dim_dict.get("id"),
                    "groupNo": dim_dict.get("group_no", dim_dict.get("groupNo")),
                    "dimensionType": dim_dict.get(
                        "dimension_type", dim_dict.get("dimensionType")
                    ),
                    "nominalValue": dim_dict.get(
                        "nominal_value", dim_dict.get("nominalValue")
                    ),
                    "toleranceValue": dim_dict.get(
                        "tolerance_value", dim_dict.get("toleranceValue")
                    ),
                    "upperTolerance": dim_dict.get(
                        "upper_tolerance", dim_dict.get("upperTolerance")
                    ),
                    "lowerTolerance": dim_dict.get(
                        "lower_tolerance", dim_dict.get("lowerTolerance")
                    ),
                    "datum": dim_dict.get("datum"),
                    "characteristic": dim_dict.get("characteristic"),
                    "notes": dim_dict.get("notes"),
                    "imageUrl": dim_dict.get("image_url", dim_dict.get("imageUrl")),
                }
            )
        return jsonify({"success": True, "data": result, "total": len(result)})

    except Exception as e:
        logger.error(f"获取项目尺寸失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取尺寸数据失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>/grouped", methods=["GET"])
def get_project_dimensions_grouped(project_id):
    """获取项目的分组尺寸数据"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        groups = service.get_dimensions_grouped(project_id)
        return jsonify({"success": True, "data": groups})

    except Exception as e:
        logger.error(f"获取分组尺寸数据失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取分组尺寸数据失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>", methods=["POST"])
def create_dimension(project_id):
    """创建新尺寸"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({"success": False, "message": "请提供尺寸数据"}), 400

        db_session = get_db_session()
        service = DimensionService(db_session)

        # 如果没有指定组号，自动分配下一个组号
        if "groupNo" not in data or not data["groupNo"]:
            data["groupNo"] = service.get_next_group_number(project_id)

        dimension = service.create_dimension(project_id, data)
        # 确保返回的数据格式与前端期望一致
        dim_dict = dimension.to_dict()
        response_data = {
            "id": dim_dict.get("id"),
            "groupNo": dim_dict.get("group_no", dim_dict.get("groupNo")),
            "dimensionType": dim_dict.get(
                "dimension_type", dim_dict.get("dimensionType")
            ),
            "nominalValue": dim_dict.get("nominal_value", dim_dict.get("nominalValue")),
            "toleranceValue": dim_dict.get(
                "tolerance_value", dim_dict.get("toleranceValue")
            ),
            "upperTolerance": dim_dict.get(
                "upper_tolerance", dim_dict.get("upperTolerance")
            ),
            "lowerTolerance": dim_dict.get(
                "lower_tolerance", dim_dict.get("lowerTolerance")
            ),
            "datum": dim_dict.get("datum"),
            "characteristic": dim_dict.get("characteristic"),
            "notes": dim_dict.get("notes"),
        }

        return (
            jsonify(
                {"success": True, "message": "尺寸创建成功", "data": response_data}
            ),
            201,
        )

    except DimensionValidationError as e:
        logger.error(f"尺寸验证失败: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 400

    except Exception as e:
        logger.error(f"创建尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"创建尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/<int:dimension_id>", methods=["GET"])
def get_dimension(dimension_id):
    """获取单个尺寸"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        dimension = service.get_dimension_by_id(dimension_id)

        if not dimension:
            return jsonify({"success": False, "message": "尺寸不存在"}), 404

        result = dimension.to_dict()
        return jsonify({"success": True, "data": result})

    except Exception as e:
        logger.error(f"获取尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"获取尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/<int:dimension_id>", methods=["PUT"])
def update_dimension(dimension_id):
    """更新尺寸"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({"success": False, "message": "请提供更新数据"}), 400

        db_session = get_db_session()
        service = DimensionService(db_session)

        dimension = service.update_dimension(dimension_id, data)

        if not dimension:
            return jsonify({"success": False, "message": "尺寸不存在"}), 404

        result = dimension.to_dict()
        return jsonify({"success": True, "data": result, "message": "尺寸更新成功"})

    except DimensionValidationError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    except Exception as e:
        logger.error(f"更新尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"更新尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/<int:dimension_id>", methods=["DELETE"])
def delete_dimension(dimension_id):
    """删除尺寸"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        success = service.delete_dimension(dimension_id)
        if not success:
            return jsonify({"success": False, "message": "尺寸不存在"}), 404

        return jsonify({"success": True, "message": "尺寸删除成功"})

    except Exception as e:
        logger.error(f"删除尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"删除尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/projects/<project_id>/bulk", methods=["POST"])
def bulk_create_dimensions(project_id):
    """批量创建尺寸"""
    try:
        data = request.get_json()

        if not data or not isinstance(data, list):
            return jsonify({"success": False, "message": "请提供尺寸数据数组"}), 400

        if len(data) == 0:
            return jsonify({"success": False, "message": "尺寸数据数组不能为空"}), 400

        db_session = get_db_session()
        service = DimensionService(db_session)

        # 验证所有尺寸数据
        for idx, dim_data in enumerate(data):
            try:
                service.validate_dimension_data(dim_data)
            except DimensionValidationError as e:
                return jsonify({
                    "success": False, 
                    "message": f"第 {idx + 1} 个尺寸验证失败: {str(e)}"
                }), 400

        # 批量创建尺寸
        try:
            dimensions = service.bulk_create_dimensions(project_id, data)
        except Exception as e:
            logger.error(f"批量创建尺寸失败: {str(e)}")
            return (
                jsonify({"success": False, "message": f"批量创建尺寸失败: {str(e)}"}),
                500,
            )

        result = [dim.to_dict() for dim in dimensions]
        return (
            jsonify(
                {
                    "success": True,
                    "data": {
                        "dimensions": result,
                        "count": len(dimensions)
                    },
                    "message": f"成功创建 {len(dimensions)} 个尺寸",
                }
            ),
            201,
        )

    except DimensionValidationError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    except Exception as e:
        logger.error(f"批量创建尺寸失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"批量创建尺寸失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>/import", methods=["POST"])
def import_dimensions(project_id):
    """从Excel/CSV文件导入尺寸数据"""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "message": "请选择文件"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"success": False, "message": "请选择文件"}), 400

        # 检查文件类型
        allowed_extensions = {"xlsx", "csv"}
        if not (
            "." in file.filename
            and file.filename.rsplit(".", 1)[1].lower() in allowed_extensions
        ):
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "不支持的文件格式，请上传 XLSX 或 CSV 格式的文件",
                    }
                ),
                400,
            )

        # 读取文件内容
        import io
        from openpyxl import load_workbook
        import csv

        dimensions_data = []
        file_extension = file.filename.rsplit(".", 1)[1].lower()

        if file_extension == "xlsx":
            # 处理Excel文件
            wb = load_workbook(filename=io.BytesIO(file.read()))
            ws = wb.active

            # 读取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)

            # 读取数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        header = headers[i]
                        # 映射表头到API字段
                        header_mapping = {
                            "组号": "groupNo",
                            "尺寸类型": "dimensionType",
                            "名义值": "nominalValue",
                            "公差值": "toleranceValue",
                            "上公差": "upperTolerance",
                            "下公差": "lowerTolerance",
                            "单位": "unit",
                            "基准": "datum",
                            "特殊特性": "characteristic",
                            "FCF符号": "fcfSymbol",
                            "FCF值": "fcfValue",
                            "FCF修饰符": "fcfModifier",
                            "FCF基准序列": "fcfDatums",
                            "备注": "notes",
                            "图片URL": "imageUrl"
                        }
                        if header in header_mapping:
                            row_data[header_mapping[header]] = value
                if row_data:
                    dimensions_data.append(row_data)

        elif file_extension == "csv":
            # 处理CSV文件
            csv_data = file.read().decode('utf-8')
            reader = csv.DictReader(csv_data.splitlines())
            
            # 映射表头到API字段
            header_mapping = {
                "组号": "groupNo",
                "尺寸类型": "dimensionType",
                "名义值": "nominalValue",
                "公差值": "toleranceValue",
                "上公差": "upperTolerance",
                "下公差": "lowerTolerance",
                "单位": "unit",
                "基准": "datum",
                "特殊特性": "characteristic",
                "FCF符号": "fcfSymbol",
                "FCF值": "fcfValue",
                "FCF修饰符": "fcfModifier",
                "FCF基准序列": "fcfDatums",
                "备注": "notes",
                "图片URL": "imageUrl"
            }
            
            for row in reader:
                row_data = {}
                for header, value in row.items():
                    if header in header_mapping:
                        row_data[header_mapping[header]] = value
                if row_data:
                    dimensions_data.append(row_data)

        if not dimensions_data:
            return jsonify({"success": False, "message": "文件中没有有效的尺寸数据"}), 400

        # 批量创建尺寸
        db_session = get_db_session()
        service = DimensionService(db_session)

        dimensions = service.bulk_create_dimensions(project_id, dimensions_data)

        result = [dim.to_dict() for dim in dimensions]
        return (
            jsonify(
                {
                    "success": True,
                    "data": result,
                    "message": f"成功导入 {len(dimensions)} 个尺寸",
                }
            ),
            201,
        )

    except DimensionValidationError as e:
        return jsonify({"success": False, "message": str(e)}), 400

    except Exception as e:
        logger.error(f"导入尺寸失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"导入尺寸失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>/next-group", methods=["GET"])
def get_next_group_number(project_id):
    """获取下一个组号"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        next_group = service.get_next_group_number(project_id)
        return jsonify({"success": True, "data": {"nextGroupNumber": next_group}})

    except Exception as e:
        logger.error(f"获取下一个组号失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取下一个组号失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>/insert", methods=["POST"])
def insert_dimension_at_position(project_id):
    """在指定位置插入尺寸，后续编号自动顺延"""
    try:
        data = request.get_json()

        if not data:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "请提供尺寸数据",
                        "details": "JSON请求体为空或格式无效",
                    }
                ),
                400,
            )

        insert_position = data.get("insertPosition")
        if insert_position is None:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "请指定插入位置",
                        "details": "insertPosition字段为必填",
                    }
                ),
                400,
            )

        db_session = get_db_session()
        service = DimensionService(db_session)

        dimension = service.insert_dimension_at_position(
            project_id, insert_position, data
        )
        # 返回格式化的数据
        dim_dict = dimension.to_dict()
        response_data = {
            "id": dim_dict.get("id"),
            "groupNo": dim_dict.get("group_no", dim_dict.get("groupNo")),
            "dimensionType": dim_dict.get(
                "dimension_type", dim_dict.get("dimensionType")
            ),
            "nominalValue": dim_dict.get("nominal_value", dim_dict.get("nominalValue")),
            "toleranceValue": dim_dict.get(
                "tolerance_value", dim_dict.get("toleranceValue")
            ),
            "upperTolerance": dim_dict.get(
                "upper_tolerance", dim_dict.get("upperTolerance")
            ),
            "lowerTolerance": dim_dict.get(
                "lower_tolerance", dim_dict.get("lowerTolerance")
            ),
            "datum": dim_dict.get("datum"),
            "characteristic": dim_dict.get("characteristic"),
            "notes": dim_dict.get("notes"),
            "imageUrl": dim_dict.get("image_url", dim_dict.get("imageUrl")),
        }

        return (
            jsonify(
                {"success": True, "message": "尺寸插入成功", "data": response_data}
            ),
            201,
        )

    except Exception as e:
        logger.error(f"插入尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"插入尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/<int:dimension_id>/delete-with-reorder", methods=["DELETE"])
def delete_dimension_with_reorder(dimension_id):
    """删除尺寸并重新排序后续编号"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        success = service.delete_dimension_with_reorder(dimension_id)
        if not success:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "尺寸不存在",
                        "details": f"尺寸ID {dimension_id} 不存在或已被删除",
                    }
                ),
                404,
            )

        return jsonify({"success": True, "message": "尺寸删除成功，编号已重新排序"})

    except Exception as e:
        logger.error(f"删除尺寸失败: {str(e)}")
        return (
            jsonify(
                {
                    "success": False,
                    "message": "删除尺寸失败",
                    "details": f"删除过程中发生错误: {str(e)}",
                }
            ),
            500,
        )


@dimensions_bp.route("/upload-image", methods=["POST"])
def upload_dimension_image():
    """上传尺寸图片"""
    try:
        if "image" not in request.files:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "请选择图片文件",
                        "details": "请求中缺少image文件字段",
                    }
                ),
                400,
            )

        file = request.files["image"]
        if file.filename == "":
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "请选择图片文件",
                        "details": "图片文件名不能为空",
                    }
                ),
                400,
            )

        # 检查文件类型
        allowed_extensions = {"png", "jpg", "jpeg", "gif", "bmp", "webp"}
        if not (
            "." in file.filename
            and file.filename.rsplit(".", 1)[1].lower() in allowed_extensions
        ):
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "不支持的图片格式，请上传 PNG、JPG、JPEG、GIF、BMP 或 WEBP 格式的图片",
                        "details": f"当前文件格式: {file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else '未知'}",
                    }
                ),
                400,
            )

        db_session = get_db_session()
        service = DimensionService(db_session)

        image_url = service.save_dimension_image(file)
        return jsonify(
            {
                "success": True,
                "message": "图片上传成功",
                "data": {"imageUrl": image_url},
            }
        )

    except Exception as e:
        logger.error(f"上传图片失败: {str(e)}")
        return jsonify({"success": False, "message": f"上传图片失败: {str(e)}"}), 500


@dimensions_bp.route("/projects/<project_id>/image-dimension", methods=["POST"])
def create_image_dimension(project_id):
    """创建图片尺寸（上传图片并创建尺寸记录）"""
    try:
        # 检查是否有图片文件
        if "image" not in request.files:
            return jsonify({"success": False, "message": "请选择图片文件"}), 400

        file = request.files["image"]
        if file.filename == "":
            return jsonify({"success": False, "message": "请选择图片文件"}), 400

        # 获取其他参数
        insert_position = request.form.get("insertPosition")
        characteristic = request.form.get("characteristic", "")
        part_id = request.form.get("partId", "")  # 获取产品编号

        if not insert_position:
            return jsonify({"success": False, "message": "请指定插入位置"}), 400

        if not characteristic.strip():
            return jsonify({"success": False, "message": "请输入特殊特性"}), 400

        # 检查文件类型
        allowed_extensions = {"png", "jpg", "jpeg", "gif", "bmp", "webp"}
        if not (
            "." in file.filename
            and file.filename.rsplit(".", 1)[1].lower() in allowed_extensions
        ):
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "不支持的图片格式，请上传 PNG、JPG、JPEG、GIF、BMP 或 WEBP 格式的图片",
                    }
                ),
                400,
            )

        db_session = get_db_session()
        service = DimensionService(db_session)

        # 上传图片并获取URL
        image_url = service.save_dimension_image(file)

        # 创建图片尺寸数据
        dimension_data = {
            "partId": part_id,  # 添加产品编号
            "dimensionType": "image_dimension",
            "nominalValue": "",  # 图片尺寸不需要名义值
            "toleranceValue": "",
            "upperTolerance": "",
            "lowerTolerance": "",
            "datum": "",
            "characteristic": characteristic.strip(),
            "notes": "图片尺寸",
            "imageUrl": image_url,
        }

        # 在指定位置插入图片尺寸
        dimension = service.insert_dimension_at_position(
            project_id, int(insert_position), dimension_data
        )
        # 返回格式化的数据
        dim_dict = dimension.to_dict()
        response_data = {
            "id": dim_dict.get("id"),
            "groupNo": dim_dict.get("group_no", dim_dict.get("groupNo")),
            "dimensionType": dim_dict.get(
                "dimension_type", dim_dict.get("dimensionType")
            ),
            "nominalValue": dim_dict.get("nominal_value", dim_dict.get("nominalValue")),
            "toleranceValue": dim_dict.get(
                "tolerance_value", dim_dict.get("toleranceValue")
            ),
            "upperTolerance": dim_dict.get(
                "upper_tolerance", dim_dict.get("upperTolerance")
            ),
            "lowerTolerance": dim_dict.get(
                "lower_tolerance", dim_dict.get("lowerTolerance")
            ),
            "datum": dim_dict.get("datum"),
            "characteristic": dim_dict.get("characteristic"),
            "notes": dim_dict.get("notes"),
            "imageUrl": dim_dict.get("image_url", dim_dict.get("imageUrl")),
        }

        return (
            jsonify(
                {"success": True, "message": "图片尺寸创建成功", "data": response_data}
            ),
            201,
        )

    except Exception as e:
        logger.error(f"创建图片尺寸失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"创建图片尺寸失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/projects/<project_id>/image-only", methods=["POST"])
def create_image_only_dimension(project_id):
    """创建纯图片尺寸（只包含图片，无其他数据）"""
    try:
        data = request.get_json()

        if not data or not data.get("imageUrl"):
            return jsonify({"success": False, "message": "请提供图片URL"}), 400

        db_session = get_db_session()
        service = DimensionService(db_session)

        # 创建纯图片尺寸
        dimension_data = {
            "dimensionType": "image",
            "imageUrl": data["imageUrl"],
            "notes": data.get("notes", "图片尺寸"),
        }

        # 如果指定了插入位置，使用插入功能
        if "insertPosition" in data:
            dimension = service.insert_dimension_at_position(
                project_id, data["insertPosition"], dimension_data
            )
        else:
            # 如果没有指定组号，自动分配下一个组号
            if "groupNo" not in dimension_data or not dimension_data["groupNo"]:
                dimension_data["groupNo"] = service.get_next_group_number(project_id)
            dimension = service.create_dimension(project_id, dimension_data)
        # 返回格式化的数据
        dim_dict = dimension.to_dict()
        response_data = {
            "id": dim_dict.get("id"),
            "groupNo": dim_dict.get("group_no", dim_dict.get("groupNo")),
            "dimensionType": dim_dict.get(
                "dimension_type", dim_dict.get("dimensionType")
            ),
            "imageUrl": dim_dict.get("image_url", dim_dict.get("imageUrl")),
            "notes": dim_dict.get("notes"),
        }

        return (
            jsonify(
                {"success": True, "message": "图片尺寸创建成功", "data": response_data}
            ),
            201,
        )

    except Exception as e:
        logger.error(f"创建图片尺寸失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"创建图片尺寸失败: {str(e)}"}),
            500,
        )


# 错误处理
@dimensions_bp.errorhandler(404)
def not_found(error):
    return (
        jsonify(
            {
                "success": False,
                "message": "接口不存在",
                "details": "请求的API端点不存在或已被移除",
            }
        ),
        404,
    )


@dimensions_bp.errorhandler(500)
def internal_error(error):
    return (
        jsonify(
            {
                "success": False,
                "message": "服务器内部错误",
                "details": "服务器处理请求时发生内部错误，请检查服务器日志获取详细信息",
            }
        ),
        500,
    )


@dimensions_bp.route("/quality/validate", methods=["POST"])
def validate_dimension():
    """验证尺寸是否合格"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "请提供验证数据"}), 400

        dimension_id = data.get("dimensionId")
        actual_value = data.get("actualValue")

        if not dimension_id or not actual_value:
            return (
                jsonify(
                    {"success": False, "message": "dimensionId和actualValue为必填字段"}
                ),
                400,
            )

        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        result = quality_service.validate_dimension(dimension_id, actual_value)
        if not result.get("success"):
            return jsonify({"success": False, "message": result.get("message")}), 400

        return jsonify({"success": True, "data": result})

    except Exception as e:
        logger.error(f"验证尺寸失败: {str(e)}")
        return jsonify({"success": False, "message": f"验证尺寸失败: {str(e)}"}), 500


@dimensions_bp.route("/quality/inspect", methods=["POST"])
def save_inspection_result():
    """保存尺寸检验结果"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "请提供检验数据"}), 400

        dimension_id = data.get("dimensionId")
        actual_value = data.get("actualValue")
        inspector = data.get("inspector")
        notes = data.get("notes")

        if not dimension_id or not actual_value:
            return (
                jsonify(
                    {"success": False, "message": "dimensionId和actualValue为必填字段"}
                ),
                400,
            )

        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        inspection = quality_service.save_inspection_result(
            dimension_id, actual_value, inspector, notes
        )
        return jsonify(
            {
                "success": True,
                "message": "检验结果保存成功",
                "data": inspection.to_dict(),
            }
        )

    except Exception as e:
        logger.error(f"保存检验结果失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"保存检验结果失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/quality/project/<project_id>/pass-rate", methods=["GET"])
def get_project_pass_rate(project_id):
    """获取项目的合格率"""
    try:
        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        pass_rate_info = quality_service.get_project_pass_rate(project_id)
        return jsonify({"success": True, "data": pass_rate_info})

    except Exception as e:
        logger.error(f"获取项目合格率失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取项目合格率失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route(
    "/quality/project/<project_id>/part/<part_id>/pass-rate", methods=["GET"]
)
def get_part_pass_rate(project_id, part_id):
    """获取零件的合格率"""
    try:
        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        pass_rate_info = quality_service.get_part_pass_rate(project_id, part_id)
        return jsonify({"success": True, "data": pass_rate_info})

    except Exception as e:
        logger.error(f"获取零件合格率失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取零件合格率失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/quality/report/generate", methods=["POST"])
def generate_quality_report():
    """生成质量报告"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "请提供报告数据"}), 400

        project_id = data.get("projectId")
        report_name = data.get("reportName")
        generated_by = data.get("generatedBy")
        notes = data.get("notes")

        if not project_id or not report_name:
            return (
                jsonify(
                    {"success": False, "message": "projectId和reportName为必填字段"}
                ),
                400,
            )

        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        report = quality_service.generate_quality_report(
            project_id, report_name, generated_by, notes
        )
        return jsonify(
            {"success": True, "message": "质量报告生成成功", "data": report.to_dict()}
        )

    except Exception as e:
        logger.error(f"生成质量报告失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"生成质量报告失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/quality/dimension/<int:dimension_id>/history", methods=["GET"])
def get_inspection_history(dimension_id):
    """获取尺寸的检验历史"""
    try:
        limit = request.args.get("limit", 10, type=int)

        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        inspections = quality_service.get_inspection_history(dimension_id, limit)
        result = [insp.to_dict() for insp in inspections]

        return jsonify({"success": True, "data": result, "total": len(result)})

    except Exception as e:
        logger.error(f"获取检验历史失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取检验历史失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/quality/project/<project_id>/inspections", methods=["GET"])
def get_project_inspections(project_id):
    """获取项目的检验结果"""
    try:
        start_date = request.args.get("startDate")
        end_date = request.args.get("endDate")

        db_session = get_db_session()
        quality_service = QualityControlService(db_session)

        inspections = quality_service.get_project_inspections(
            project_id, start_date, end_date
        )
        result = [insp.to_dict() for insp in inspections]

        return jsonify({"success": True, "data": result, "total": len(result)})

    except Exception as e:
        logger.error(f"获取项目检验结果失败: {str(e)}")
        return (
            jsonify({"success": False, "message": f"获取项目检验结果失败: {str(e)}"}),
            500,
        )


@dimensions_bp.route("/export/excel", methods=["GET"])
def export_dimensions_excel():
    """导出尺寸数据为Excel格式"""
    try:
        project_id = request.args.get("projectId")
        part_id = request.args.get("partId")

        if not project_id:
            return jsonify({"success": False, "message": "请提供projectId参数"}), 400

        db_session = get_db_session()
        service = DimensionService(db_session)

        # 获取尺寸数据
        if part_id:
            dimensions = service.get_dimensions_by_part_number(project_id, part_id)
        else:
            dimensions = service.get_dimensions_by_project(project_id)

        # 生成Excel文件
        import io
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "尺寸数据"

        # 设置表头
        headers = [
            "组号", "尺寸类型", "名义值", "公差值", "上公差", "下公差", 
            "单位", "基准", "特殊特性", "FCF符号", "FCF值", "FCF修饰符", 
            "FCF基准序列", "备注", "图片URL"
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

        # 写入数据
        for row_idx, dim in enumerate(dimensions, 2):
            dim_dict = dim.to_dict()
            ws.cell(row=row_idx, column=1, value=dim_dict.get("groupNo"))
            ws.cell(row=row_idx, column=2, value=dim_dict.get("dimensionType"))
            ws.cell(row=row_idx, column=3, value=dim_dict.get("nominalValue"))
            ws.cell(row=row_idx, column=4, value=dim_dict.get("toleranceValue"))
            ws.cell(row=row_idx, column=5, value=dim_dict.get("upperTolerance"))
            ws.cell(row=row_idx, column=6, value=dim_dict.get("lowerTolerance"))
            ws.cell(row=row_idx, column=7, value=dim_dict.get("unit"))
            ws.cell(row=row_idx, column=8, value=dim_dict.get("datum"))
            ws.cell(row=row_idx, column=9, value=dim_dict.get("characteristic"))
            ws.cell(row=row_idx, column=10, value=dim_dict.get("fcfSymbol"))
            ws.cell(row=row_idx, column=11, value=dim_dict.get("fcfValue"))
            ws.cell(row=row_idx, column=12, value=dim_dict.get("fcfModifier"))
            ws.cell(row=row_idx, column=13, value=dim_dict.get("fcfDatums"))
            ws.cell(row=row_idx, column=14, value=dim_dict.get("notes"))
            ws.cell(row=row_idx, column=15, value=dim_dict.get("imageUrl"))

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        # 发送文件
        filename = f"尺寸数据_{project_id}_{part_id if part_id else '全部'}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"导出Excel失败: {str(e)}")
        return jsonify({"success": False, "message": f"导出Excel失败: {str(e)}"}), 500


@dimensions_bp.route("/clear-all", methods=["POST"])
def clear_all_dimensions():
    """清空所有尺寸数据，包括虚拟数据"""
    try:
        db_session = get_db_session()
        service = DimensionService(db_session)

        deleted_count = service.clear_all_dimensions()
        return jsonify({
            "success": True,
            "message": f"成功清空所有尺寸数据，共删除 {deleted_count} 条记录"
        })

    except Exception as e:
        logger.error(f"清空尺寸数据失败: {str(e)}")
        return jsonify({"success": False, "message": f"清空尺寸数据失败: {str(e)}"}), 500
