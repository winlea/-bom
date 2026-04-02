#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多尺寸验证工具 - 测试一个零件编号对应多个尺寸的情况
验证图纸尺寸从B12开始嵌入到表格的功能
"""

import json
import os
import random
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


class MultiDimensionValidator:
    """多尺寸验证器"""

    def __init__(self):
        self.test_results = []
        self.output_dir = "validation_output"
        os.makedirs(self.output_dir, exist_ok=True)

    def create_test_part_with_multiple_dimensions(
        self, part_number, dimensions_count=5
    ):
        """创建包含多个尺寸的测试零件"""
        dimensions = []
        for i in range(dimensions_count):
            dimensions.append(
                {
                    "name": f"DIM_{i+1}",
                    "value": round(random.uniform(10.0, 200.0), 2),
                    "tolerance": f"±{round(random.uniform(0.1, 0.5), 2)}",
                    "unit": "mm",
                }
            )

        return {
            "part_number": part_number,
            "part_name": f"测试零件_{part_number}",
            "type": "MULTI_DIMENSION_TEST",
            "dimensions": dimensions,
            "total_dimensions": len(dimensions),
        }

    def create_template_with_multiple_dimensions(self, part_data, template_name):
        """创建包含多个尺寸的模板文件"""
        template_path = os.path.join(self.output_dir, template_name)

        # 创建新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "零件信息"

        # 设置标题
        ws["A1"] = f"零件信息 - {part_data['part_number']}"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")

        # 基本信息
        ws["A3"] = "零件编号"
        ws["B3"] = part_data["part_number"]
        ws["A4"] = "零件名称"
        ws["B4"] = part_data["part_name"]
        ws["A5"] = "类型"
        ws["B5"] = part_data["type"]

        # 尺寸标题
        ws["A7"] = "尺寸信息"
        ws["A7"].font = Font(size=14, bold=True)

        # 从B12开始嵌入尺寸
        start_row = 12
        ws["B11"] = "图纸尺寸"
        ws["B11"].font = Font(bold=True)

        for idx, dim in enumerate(part_data["dimensions"]):
            row = start_row + idx

            # 尺寸名称
            ws[f"B{row}"] = dim["name"]
            ws[f"C{row}"] = str(dim["value"])
            ws[f"D{row}"] = dim["unit"]
            ws[f"E{row}"] = dim["tolerance"]

            # 设置样式
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].fill = PatternFill(
                    start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
                )

        # 设置列宽
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 15

        wb.save(template_path)
        return template_path

    def validate_dimension_placement(self, template_path, expected_count):
        """验证尺寸是否正确放置在B12开始的区域"""
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            actual_count = 0
            start_row = 12

            # 检查B12开始的区域
            for row in range(start_row, start_row + expected_count + 5):  # 多检查几行
                cell_value = ws[f"B{row}"].value
                if cell_value and str(cell_value).startswith("DIM_"):
                    actual_count += 1

            # 检查尺寸值是否正确
            dimension_values = []
            for row in range(start_row, start_row + expected_count):
                name = ws[f"B{row}"].value
                value = ws[f"C{row}"].value
                if name and value:
                    dimension_values.append({"name": str(name), "value": str(value)})

            return {
                "success": actual_count == expected_count,
                "expected_count": expected_count,
                "actual_count": actual_count,
                "dimension_values": dimension_values,
                "placement_check": (
                    "B12区域放置正确"
                    if actual_count == expected_count
                    else "B12区域放置不正确"
                ),
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "placement_check": "文件读取失败",
            }

    def run_multi_dimension_test(self):
        """运行多尺寸测试"""
        print("开始多尺寸验证测试...")

        test_cases = [
            {"part_number": "4266007", "dimensions_count": 3},
            {"part_number": "4266008", "dimensions_count": 5},
            {"part_number": "4266009", "dimensions_count": 8},
            {"part_number": "4266010", "dimensions_count": 10},
        ]

        for case in test_cases:
            print(
                f"测试零件 {case['part_number']} 的 {case['dimensions_count']} 个尺寸..."
            )

            # 创建测试零件
            test_part = self.create_test_part_with_multiple_dimensions(
                case["part_number"], case["dimensions_count"]
            )

            # 创建模板
            template_name = f"multi_dim_{case['part_number']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            template_path = self.create_template_with_multiple_dimensions(
                test_part, template_name
            )

            # 验证尺寸放置
            validation_result = self.validate_dimension_placement(
                template_path, case["dimensions_count"]
            )

            # 记录结果
            result = {
                "part_number": case["part_number"],
                "dimensions_count": case["dimensions_count"],
                "template_file": template_name,
                "template_path": template_path,
                "validation": validation_result,
                "test_timestamp": datetime.now().isoformat(),
            }

            self.test_results.append(result)

            status = "✓ 通过" if validation_result["success"] else "✗ 失败"
            print(
                f"{status} 零件 {case['part_number']} - 找到 {validation_result['actual_count']} 个尺寸"
            )

        # 生成测试报告
        self.generate_test_report()

    def generate_test_report(self):
        """生成测试报告"""
        report = {
            "test_name": "多尺寸验证测试",
            "test_date": datetime.now().isoformat(),
            "total_tests": len(self.test_results),
            "passed_tests": sum(
                1 for r in self.test_results if r["validation"]["success"]
            ),
            "failed_tests": sum(
                1 for r in self.test_results if not r["validation"]["success"]
            ),
            "results": self.test_results,
        }

        report_path = os.path.join(self.output_dir, "multi_dimension_test_report.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)

        print(f"\n测试报告已保存: {report_path}")

        # 打印总结
        print(f"\n=== 测试总结 ===")
        print(f"总测试数: {report['total_tests']}")
        print(f"通过测试: {report['passed_tests']}")
        print(f"失败测试: {report['failed_tests']}")

        if report["failed_tests"] > 0:
            print("\n失败详情:")
            for result in self.test_results:
                if not result["validation"]["success"]:
                    print(
                        f"- 零件 {result['part_number']}: {result['validation']['placement_check']}"
                    )

    def run_quick_test(self):
        """运行快速测试"""
        print("运行快速多尺寸测试...")

        # 创建一个测试零件，包含4个尺寸
        test_part = self.create_test_part_with_multiple_dimensions("TEST_001", 4)

        # 创建模板
        template_name = "quick_multi_dim_test.xlsx"
        template_path = self.create_template_with_multiple_dimensions(
            test_part, template_name
        )

        # 验证
        result = self.validate_dimension_placement(template_path, 4)

        print(f"快速测试结果: {'通过' if result['success'] else '失败'}")
        print(f"找到尺寸数量: {result['actual_count']}")

        return result


if __name__ == "__main__":
    validator = MultiDimensionValidator()

    if len(os.sys.argv) > 1 and os.sys.argv[1] == "--quick":
        validator.run_quick_test()
    else:
        validator.run_multi_dimension_test()
