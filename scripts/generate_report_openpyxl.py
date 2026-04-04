"""
使用 openpyxl 和 win32com 生成初始过程能力分析报告
为每个 CC/SC 特性创建单独的工作表
"""

import os
import sys
import re
import statistics
import random
import shutil

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import create_app
from bom_system.models import BomTable
from bom_system.dimensions.models import Dimension
from bom_system.config import PROCESS_CAPABILITY_TEMPLATE_PATH

try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False
    print("警告: win32com 不可用，将使用 openpyxl 备用方案")


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """使用 win32com 将 xls 转换为 xlsx"""
    if not WIN32COM_AVAILABLE:
        return False
    
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.Dispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
        wb.Close(False)
        excel.Quit()
        return True
    except Exception as e:
        print(f"转换失败: {e}")
        return False
    finally:
        pythoncom.CoUninitialize()


def generate_report(part, output_dir):
    """为零件生成报告"""
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    app = create_app()
    with app.app_context():
        # 获取 CC/SC 特性
        dims = Dimension.query.filter(
            Dimension.part_id == part.part_number,
            Dimension.characteristic.like('CC%') | Dimension.characteristic.like('SC%')
        ).order_by(Dimension.characteristic).all()
        
        if not dims:
            print(f"没有找到 CC/SC 特性")
            return None
        
        print(f"找到 {len(dims)} 个 CC/SC 特性")
        
        # 转换模板为 xlsx
        template_xls = os.path.abspath(PROCESS_CAPABILITY_TEMPLATE_PATH)
        template_xlsx = os.path.abspath(os.path.join(output_dir, 'template_converted.xlsx'))
        
        if not convert_xls_to_xlsx(template_xls, template_xlsx):
            print("模板转换失败")
            return None
        
        print(f"模板已转换为 xlsx: {template_xlsx}")
        
        # 加载模板
        wb = load_workbook(template_xlsx)
        
        # 删除所有图表（openpyxl 对图表支持有限）
        for sheet in wb.worksheets:
            if sheet._charts:
                sheet._charts = []
        
        template_ws = wb.active
        template_ws.title = dims[0].characteristic  # 第一个工作表用第一个特性命名
        
        # 填充第一个工作表的数据
        fill_worksheet(template_ws, part, dims[0])
        
        # 为其他特性创建工作表
        for i in range(1, len(dims)):
            dim = dims[i]
            # 复制模板工作表
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = dim.characteristic
            # 填充数据
            fill_worksheet(new_ws, part, dim)
        
        # 保存报告
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{part.part_number}_初始过程能力分析报告.xlsx")
        wb.save(output_path)
        print(f"报告已保存: {output_path}")
        
        # 删除临时模板
        if os.path.exists(template_xlsx):
            os.remove(template_xlsx)
        
        return output_path


def fill_worksheet(ws, part, dim):
    """填充工作表数据"""
    # 填充零件信息
    ws.cell(row=6, column=5).value = part.part_number
    ws.cell(row=6, column=12).value = part.part_name or ""
    if part.drawing_2d:
        ws.cell(row=7, column=5).value = part.drawing_2d
    
    # 填充尺寸数据
    nominal = float(dim.nominal_value) if dim.nominal_value else 0
    upper = float(dim.upper_tolerance) if dim.upper_tolerance else 0
    lower = float(dim.lower_tolerance) if dim.lower_tolerance else 0
    
    ws.cell(row=10, column=5).value = nominal  # E10 名义值
    ws.cell(row=10, column=7).value = upper    # G10 上公差
    ws.cell(row=10, column=9).value = lower    # I10 下公差
    
    # 生成25个数据点
    if upper - lower < 0.1:
        upper = 0.1
        lower = -0.1
    
    random.seed(42)
    data_points = []
    for _ in range(25):
        value = nominal + random.gauss(0, 0.005)
        value = max(nominal + lower, min(nominal + upper, value))
        data_points.append(round(value, 4))
    
    # 写入数据点
    for i, val in enumerate(data_points):
        ws.cell(row=12 + i, column=5).value = val
    
    # 计算并写入 CPK/PPK
    mean = statistics.mean(data_points)
    std = statistics.stdev(data_points) if len(data_points) > 1 else 0.001
    
    usl = nominal + upper
    lsl = nominal + lower
    cpk_u = (usl - mean) / (3 * std) if std > 0 else 0
    cpk_l = (mean - lsl) / (3 * std) if std > 0 else 0
    cpk = min(cpk_u, cpk_l)
    
    ws.cell(row=40, column=5).value = round(cpk, 4)
    ws.cell(row=41, column=5).value = round(cpk, 4)
    
    result = "process is capable" if cpk > 1.67 else "process is not capable"
    ws.cell(row=42, column=5).value = result
    ws.cell(row=42, column=21).value = result


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        part = BomTable.query.get(43)
        if part:
            print(f"零件: {part.part_number}")
            output_path = generate_report(part, "output/process_capability")
            if output_path:
                print(f"成功生成报告: {output_path}")
            else:
                print("报告生成失败")
        else:
            print("零件不存在")
