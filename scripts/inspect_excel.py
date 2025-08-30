import sys

from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: python scripts/inspect_excel.py <path_to_xlsx>")
    sys.exit(1)

path = sys.argv[1]
wb = load_workbook(path, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
if not rows:
    print("Empty workbook")
    sys.exit(0)

headers = [str(v or "").strip() for v in rows[0]]
print("Headers:")
print(headers)

print("Sample rows:")
for r in rows[1:6]:
    print([str(v) if v is not None else "" for v in r])
