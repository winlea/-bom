from openpyxl import load_workbook

# 加载ODS文件
wb = load_workbook('ODS_test.xlsx')
sheet = wb.active

# 打印表头
print('表头:')
for row in sheet.iter_rows(min_row=4, max_row=4, values_only=True):
    print(row)

# 打印数据
print('\n数据:')
for row in sheet.iter_rows(min_row=5, values_only=True):
    print(row)
